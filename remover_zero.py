"""
Passo 1: em cada aba, remove da tabela de OBs (colunas "OB"/"VALOR") todas
as linhas em que VALOR = 0.

Pode ser rodado isoladamente (ver bloco __main__ no final), embora
processar_pagamentos.py já chame essa mesma etapa automaticamente antes da
conciliação - não é necessário rodar os dois separadamente.

Uso:
    python remover_zero.py caminho_da_planilha.xlsx
"""

import argparse

import openpyxl

from processar_pagamentos import (
    NORMAL_FONT, BORDER, NO_BORDER, NO_FILL, CENTER, RIGHT,
    localizar_bloco_ob, _ler_linhas_ob, _normalizar_abas,
)


def remover_valor_zero_aba(ws):
    """Remove as OBs com VALOR = 0 dentro dessa aba. Devolve
    (removidas, restantes) ou None se a aba não tiver bloco OB/VALOR."""
    col_ob, col_valor, linha_header = localizar_bloco_ob(ws)
    if col_ob is None:
        return None

    linha_inicial = linha_header + 1
    linhas = _ler_linhas_ob(ws, col_ob, col_valor, linha_inicial)

    total_antes = len(linhas)
    linhas_filtradas = [(ob, valor) for ob, valor in linhas if valor != 0]
    removidas = total_antes - len(linhas_filtradas)

    # limpa toda a área antiga do bloco (valor, borda e preenchimento)
    ultima_linha_bloco = linha_inicial + total_antes
    for r in range(linha_inicial, max(ws.max_row, ultima_linha_bloco) + 1):
        for col in (col_ob, col_valor):
            cell = ws.cell(row=r, column=col)
            cell.value = None
            cell.border = NO_BORDER
            cell.fill = NO_FILL

    # reescreve só as linhas que sobraram
    for i, (ob, valor) in enumerate(linhas_filtradas):
        r = linha_inicial + i
        c_ob = ws.cell(row=r, column=col_ob, value=ob)
        c_ob.font = NORMAL_FONT
        c_ob.border = BORDER
        c_ob.alignment = CENTER

        c_valor = ws.cell(row=r, column=col_valor, value=valor)
        c_valor.font = NORMAL_FONT
        c_valor.border = BORDER
        c_valor.number_format = "#,##0.00"
        c_valor.alignment = RIGHT

    return removidas, len(linhas_filtradas)


def remover_valor_zero(caminho, abas=None):
    wb = openpyxl.load_workbook(caminho)
    abas_selecionadas = _normalizar_abas(abas)
    abas_alvo = abas_selecionadas or wb.sheetnames

    total_removidas = 0
    abas_processadas = 0
    for nome in abas_alvo:
        if nome not in wb.sheetnames:
            print(f"Aba '{nome}' não encontrada. Ignorando.")
            continue

        ws = wb[nome]
        resultado = remover_valor_zero_aba(ws)
        if resultado is None:
            continue
        removidas, restantes = resultado
        abas_processadas += 1
        total_removidas += removidas
        print(f"[{nome}] OBs com VALOR = 0 removidas: {removidas} | restantes: {restantes}")

    wb.save(caminho)

    if abas_processadas == 0:
        print("Nenhuma aba com tabela OB/VALOR foi encontrada. Nada foi alterado.")
    else:
        print(f"\nTotal de OBs removidas: {total_removidas}, em {abas_processadas} aba(s).")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Remove OBs com VALOR = 0 de todas as abas de uma planilha.")
    parser.add_argument("caminho", nargs="?", default="conciliacao.xlsx", help="Caminho da planilha Excel")
    parser.add_argument("--aba", "--abas", dest="abas", action="append", help="Nome da aba a ser processada. Pode ser informado mais de uma vez.")
    args = parser.parse_args()

    remover_valor_zero(args.caminho, abas=args.abas)