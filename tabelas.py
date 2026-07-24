#!/usr/bin/env python3
"""
Busca campos específicos dentro de uma planilha xlsx (ex.: Formulário de
Qualidade - FQ), localizando automaticamente a linha de cabeçalho da tabela
e lendo os valores da(s) linha(s) de dados logo abaixo.

Uso:
    python buscar_campos_xlsx.py caminho/do/arquivo.xlsx
"""

import sys
import re
import unicodedata
from difflib import SequenceMatcher

import openpyxl

# ---------------------------------------------------------------------------
# Campos que devem ser buscados na planilha.
# Adicione, remova ou edite livremente.
# ---------------------------------------------------------------------------
CAMPOS_PROCURADOS = [
    "Unidade de Medida",
    "Quantidade do Item",
    "Valor Preço Unitário",
    "UOR",
    "Conta Contábil",
]


# ---------------------------------------------------------------------------
# Utilidades de normalização e comparação "aproximada" de texto
# (mesma lógica usada no script de busca em PDF)
# ---------------------------------------------------------------------------
def normalizar(texto: str) -> str:
    if not texto:
        return ""
    texto = re.sub(r"\([^)]*\)", " ", texto)
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = texto.lower()
    texto = re.sub(r"[^a-z0-9\s]", " ", texto)
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto


def parecido(a: str, b: str, limite: float = 0.72) -> bool:
    """Compara dois textos de forma tolerante a pequenas variações de escrita.

    Mesma correção de pdfs.py:
    (1) "contém" só conta para textos com pelo menos 4 caracteres normalizados,
        para não deixar células de 1-2 letras (lixo comum de planilha/PDF)
        casarem com qualquer campo mais longo.
    (2) mesmo com 4+ caracteres, uma palavra genérica de cabeçalho de coluna
        (ex.: "ITEM") não deve vencer só por estar contida em um campo
        composto bem mais específico (ex.: "Código do Item") — só conta se o
        texto curto tiver 2+ palavras ou cobrir metade do texto longo.
    """
    na, nb = normalizar(a), normalizar(b)
    if not na or not nb:
        return False
    if na == nb:
        return True
    curto, longo = (na, nb) if len(na) <= len(nb) else (nb, na)
    if len(curto) >= 4 and curto in longo:
        if len(curto.split()) >= 2 or len(curto) / len(longo) >= 0.5:
            return True
    return SequenceMatcher(None, na, nb).ratio() >= limite


def valor_vazio(valor) -> bool:
    """True se a célula estiver vazia ou contiver só espaços em branco."""
    return valor is None or (isinstance(valor, str) and not valor.strip())


# ---------------------------------------------------------------------------
# Localização da tabela dentro da planilha
# ---------------------------------------------------------------------------
def encontrar_linha_cabecalho(ws, campos_procurados):
    """
    Percorre as linhas da planilha e devolve a que mais parece um cabeçalho
    de tabela: linha com várias células de texto que batem (aproximadamente)
    com os nomes dos campos procurados.
    """
    melhor_linha = None
    melhor_pontuacao = 0

    for linha in ws.iter_rows(min_row=1, max_row=ws.max_row):
        celulas_texto = [c for c in linha if isinstance(c.value, str) and c.value.strip()]
        if len(celulas_texto) < 3:
            continue
        pontuacao = sum(
            1
            for celula in celulas_texto
            for campo in campos_procurados
            if parecido(celula.value, campo)
        )
        if pontuacao > melhor_pontuacao:
            melhor_pontuacao = pontuacao
            melhor_linha = linha

    return melhor_linha, melhor_pontuacao


def mapear_colunas(linha_cabecalho):
    """Devolve {indice_da_coluna: texto_do_cabecalho} para a linha dada."""
    return {
        celula.column: celula.value.strip()
        for celula in linha_cabecalho
        if isinstance(celula.value, str) and celula.value.strip()
    }


def extrair_linhas_de_dados(ws, linha_cabecalho, colunas):
    """
    A partir da linha seguinte ao cabeçalho, coleta as linhas de dados da
    tabela. Pula linhas em branco/whitespace (comuns como separador visual)
    e para ao encontrar uma linha de total (ex.: 'Total') ou o fim da tabela.
    """
    linha_inicio = linha_cabecalho[0].row + 1
    linhas_dados = []

    for num_linha in range(linha_inicio, ws.max_row + 1):
        valores = {col: ws.cell(row=num_linha, column=col).value for col in colunas}

        primeira_col = min(colunas.keys())
        primeira_celula = valores.get(primeira_col)

        # Linha de total encerra a tabela
        if isinstance(primeira_celula, str) and parecido(primeira_celula, "Total"):
            break

        # Linha totalmente em branco: pula (pode ser um separador visual)
        if all(valor_vazio(v) for v in valores.values()):
            continue

        linhas_dados.append(valores)

    return linhas_dados


def buscar_campo(campo, colunas, linhas_dados):
    """Encontra a coluna cujo cabeçalho é parecido com `campo` e devolve os
    valores dessa coluna em cada linha de dados."""
    coluna_encontrada = None
    for col_idx, texto_cabecalho in colunas.items():
        if parecido(texto_cabecalho, campo):
            coluna_encontrada = col_idx
            break

    if coluna_encontrada is None:
        return None

    valores = []
    for linha in linhas_dados:
        v = linha.get(coluna_encontrada)
        if isinstance(v, str):
            v = v.strip()
        valores.append(v)

    return valores


# ---------------------------------------------------------------------------
# Pós-processamento específico de alguns campos: a célula tem mais texto do
# que o valor que interessa, então depois de achar a coluna certa a gente
# ainda recorta o valor final.
# ---------------------------------------------------------------------------
def _extrair_sigla_entre_parenteses(valor):
    """
    'Local para Faturamento' vem como 'BRASILIA - MANUTENCAO BRA(M)', mas só
    interessa a sigla final entre parênteses, ex.: 'BRA(M)'. Pega o último
    token colado a um '(...)' no fim do texto; se não achar esse padrão,
    devolve o valor original (sem quebrar quando o formato for diferente).
    """
    if not isinstance(valor, str):
        return valor
    match = re.search(r"(\S+\([^)]+\))\s*$", valor)
    return match.group(1) if match else valor


PROCESSAMENTO_ESPECIAL = {
    "Local para Faturamento": _extrair_sigla_entre_parenteses,
}


def aplicar_processamento_especial(campo, valores):
    """Se `campo` bater (aproximadamente) com uma chave de
    PROCESSAMENTO_ESPECIAL, aplica a função de recorte em cada valor da
    lista. Caso contrário, devolve os valores como vieram da planilha."""
    if valores is None:
        return None
    for chave, funcao in PROCESSAMENTO_ESPECIAL.items():
        if parecido(chave, campo):
            return [funcao(v) if v is not None else v for v in valores]
    return valores


# ---------------------------------------------------------------------------
# Função reutilizável (chamada pelo main.py e também usável via CLI)
# ---------------------------------------------------------------------------
def processar_xlsx(caminho_xlsx: str, campos_procurados: list = None) -> dict:
    """
    Analisa a planilha em `caminho_xlsx` e devolve um dicionário
    {campo: lista_de_valores_ou_None}, um valor por linha de dados da tabela
    encontrada na primeira aba onde ela aparecer.
    Lança exceção se o arquivo não puder ser aberto/lido.

    `campos_procurados` permite passar uma lista de campos específica
    (ex.: a combinação fluxo+documento decidida em main.py/doc_types.py)
    em vez da lista fixa CAMPOS_PROCURADOS. Se omitido, usa a lista fixa
    (mantém o comportamento antigo para quem chama/roda este arquivo
    isoladamente).
    """
    if campos_procurados is None:
        campos_procurados = CAMPOS_PROCURADOS

    wb = openpyxl.load_workbook(caminho_xlsx, data_only=True)

    for ws in wb.worksheets:
        linha_cabecalho, pontuacao = encontrar_linha_cabecalho(ws, campos_procurados)
        if not linha_cabecalho or pontuacao == 0:
            continue  # esta aba não parece ter a tabela procurada

        colunas = mapear_colunas(linha_cabecalho)
        linhas_dados = extrair_linhas_de_dados(ws, linha_cabecalho, colunas)

        resultado = {
            "_aba": ws.title,
            "_linha_cabecalho": linha_cabecalho[0].row,
            "_num_linhas_dados": len(linhas_dados),
        }
        for campo in campos_procurados:
            valores = buscar_campo(campo, colunas, linhas_dados)
            resultado[campo] = aplicar_processamento_especial(campo, valores)

        return resultado

    return None  # nenhuma tabela reconhecível em nenhuma aba


def extrair_texto_xlsx(caminho_xlsx: str) -> str:
    """
    Concatena o conteúdo textual da planilha (nomes das abas + valores de
    texto de todas as células) em uma única string. Usado por
    doc_types.identificar_documento() para reconhecer o tipo de documento
    pelo conteúdo, do mesmo jeito que pdfs.extrair_texto() faz para PDFs.
    """
    partes = []
    wb = openpyxl.load_workbook(caminho_xlsx, data_only=True)
    for ws in wb.worksheets:
        partes.append(ws.title)
        for linha in ws.iter_rows():
            for celula in linha:
                if isinstance(celula.value, str) and celula.value.strip():
                    partes.append(celula.value.strip())
    return "\n".join(partes)


def imprimir_resultado(caminho_xlsx: str, resultado: dict) -> None:
    print(f"Planilha analisada: {caminho_xlsx}\n")

    if resultado is None:
        print("[ERRO] Não foi encontrada nenhuma tabela reconhecível na planilha.")
        return

    print(f"[Aba: {resultado['_aba']}] tabela encontrada na linha {resultado['_linha_cabecalho']}, "
          f"{resultado['_num_linhas_dados']} linha(s) de dados.\n")

    for campo, valores in resultado.items():
        if campo.startswith("_"):
            continue
        if valores is None:
            print(f"[ERRO] Não foi possível encontrar o campo '{campo}' na planilha.")
        elif len(valores) == 1:
            print(f"[SUCESSO] Campo '{campo}' encontrado. Valor: {valores[0]}")
        else:
            print(f"[SUCESSO] Campo '{campo}' encontrado ({len(valores)} linhas). "
                  f"Valores: {valores}")


# ---------------------------------------------------------------------------
# Programa principal (uso via linha de comando, standalone)
# ---------------------------------------------------------------------------
def main():
    if len(sys.argv) < 2:
        print("Uso: python tabelas.py caminho/do/arquivo.xlsx")
        sys.exit(1)

    caminho_xlsx = sys.argv[1]

    try:
        resultado = processar_xlsx(caminho_xlsx)
    except Exception as erro:
        print(f"[ERRO] Não foi possível abrir o arquivo '{caminho_xlsx}': {erro}")
        sys.exit(1)

    imprimir_resultado(caminho_xlsx, resultado)


if __name__ == "__main__":
    main()