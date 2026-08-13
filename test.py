"""
Script pra TESTAR o pipeline de conciliação (Passos 2 a 4) na planilha
"conciliacao-julho-2026.xlsx", cruzando com os extratos de "saida.xlsx",
rodando UMA ABA POR VEZ - em vez de processar a planilha inteira de uma
tacada só, como fazem os scripts originais.

O que ele faz, para cada aba:
    Passo 2 - separar_tipo.py  -> separa Recebimentos/Pagamentos
    Passo 3 - remover_zero.py  -> remove OBs com VALOR = 0
    Passo 4 - processar_pagamentos.py (com --extratos)
              -> concilia pagamento <-> OB e cruza com saida.xlsx

Por padrão:
    - Edita a própria planilha "conciliaca-julho-2026.xlsx" (sem criar
      cópia). Se você quiser testar sem mexer no arquivo original, use
      --copia (aí ele trabalha em "..._TESTE.xlsx" e não toca no real).
    - Pausa depois de cada aba, esperando você apertar ENTER pra seguir
      pra próxima (dá pra abrir o Excel e conferir o resultado antes de
      continuar). Digite 'q' na pausa pra parar por ali.

Uso:
    # roda TODAS as abas, uma de cada vez, com pausa entre elas
    python test.py

    # roda só uma aba específica (ou várias, repetindo --aba)
    python test.py --aba "07.07.26"
    python test.py --aba "07.07.26" --aba "08.07.26"

    # roda tudo sem pausar entre abas
    python test.py --sem-pausa

    # usa outros arquivos / outro tamanho de combinação
    python test.py --conciliacao outra.xlsx --extratos extratos.xlsx --max-combinacao-pagamentos 8

    # trabalha numa cópia de teste ("..._TESTE.xlsx"), sem tocar no original
    python test.py --copia

    python test.py --sem-pausa --aba "01.07.26" --aba "02.07.26" --aba "03.07.26" --aba "04.07.26" --aba "05.07.26" --aba "06.07.26" --aba "07.07.26" --aba "08.07.26" --aba "09.07.26" --aba "10.07.26" --aba "11.07.26" --aba "12.07.26" --aba "13.07.26" --aba "14.07.26" --aba "15.07.26" --aba "16.07.26" --aba "17.07.26" --aba "18.07.26" --aba "19.07.26" --aba "20.07.26" --aba "21.07.26" --aba "22.07.26" --aba "23.07.26" --aba "24.07.26" --aba "25.07.26" --aba "26.07.26" --aba "27.07.26" --aba "28.07.26" --aba "29.07.26" --aba "30.07.26" --aba "31.07.26"
    python test.py --sem-pausa --aba "01.07.26" --aba "02.07.26" --aba "03.07.26" --aba "04.07.26" --aba "05.07.26" --aba "06.07.26" --aba "08.07.26" --aba "09.07.26" --aba "11.07.26" --aba "12.07.26" --aba "13.07.26" --aba "14.07.26" --aba "15.07.26" --aba "16.07.26" --aba "17.07.26" --aba "18.07.26" --aba "19.07.26" --aba "20.07.26" --aba "21.07.26" --aba "22.07.26" --aba "23.07.26" --aba "24.07.26" --aba "25.07.26" --aba "26.07.26" --aba "27.07.26" --aba "28.07.26" --aba "29.07.26" --aba "30.07.26" --aba "31.07.26"
"""

import argparse
import shutil
import sys
import traceback
from pathlib import Path

import openpyxl

from separar_tipo import processar_aba
from remover_zero import remover_valor_zero
from extratos import atualizar_status_sem_conciliacao


def separar_tipo_aba(caminho, aba):
    """Roda o Passo 2 (separar Recebimentos/Pagamentos) só em UMA aba.
    Os outros passos (remover_zero, atualizar_status_sem_conciliacao) já
    aceitam `abas=[...]` nativamente; só o separar_tipo.py original não
    tinha essa opção, então replicamos aqui o mesmo padrão."""
    wb = openpyxl.load_workbook(caminho)
    if aba not in wb.sheetnames:
        print(f"  Aba '{aba}' não encontrada em {caminho}. Ignorando.")
        return

    ws = wb[aba]
    resultado = processar_aba(ws, wb.epoch)
    wb.save(caminho)

    if resultado is None:
        print(f"  [Passo 2] '{aba}' pulada (sem coluna 'Tipo' no cabeçalho - aba provavelmente não é de lançamentos diários).")
        return

    extra_txt = f", extra OB/VALOR: {resultado['extra']}" if resultado["extra"] is not None else ""
    print(f"  [Passo 2] Recebimentos: {resultado['recebimentos']} | "
          f"Pagamentos: {resultado['pagamentos']} | Outros: {resultado['outros']}{extra_txt}")
    if resultado["avisos"]:
        print(f"    ATENÇÃO: {len(resultado['avisos'])} célula(s) fora do padrão conhecido (não foram movidas):")
        for r, col_letra, v in resultado["avisos"][:10]:
            print(f"      - {col_letra}{r}: {v!r}")


def rodar_aba(caminho_conciliacao, caminho_extratos, aba, max_combinacao):
    print(f"\n{'=' * 70}\nABA: {aba}\n{'=' * 70}")

    print("-- Passo 2: separar Recebimentos/Pagamentos --")
    separar_tipo_aba(caminho_conciliacao, aba)

    print("\n-- Passo 3: remover OBs com VALOR = 0 --")
    remover_valor_zero(caminho_conciliacao, abas=[aba])

    print("\n-- Passo 4: conciliar pagamento <-> OB + cruzar com extratos --")
    atualizar_status_sem_conciliacao(
        caminho_conciliacao, caminho_extratos, max_combinacao=max_combinacao, abas=[aba]
    )


def main():
    parser = argparse.ArgumentParser(
        description="Testa o pipeline de conciliação (Passos 2 a 4) aba por aba."
    )
    parser.add_argument("--conciliacao", default="conciliacao-julho-2026.xlsx",
                         help="Planilha de conciliação (padrão: conciliaca-julho-2026.xlsx)")
    parser.add_argument("--extratos", default="saida.xlsx",
                         help="Planilha de extratos (padrão: saida.xlsx)")
    parser.add_argument("--aba", "--abas", dest="abas", action="append",
                         help="Nome de uma aba a testar. Pode repetir a flag pra mais de uma. "
                              "Se omitido, roda TODAS as abas da planilha, uma por vez.")
    parser.add_argument("--sem-pausa", action="store_true",
                         help="Não pausa entre as abas (roda tudo direto, sem pedir ENTER).")
    parser.add_argument("--copia", action="store_true",
                         help="Em vez de editar a planilha original, cria e usa uma cópia "
                              "('..._TESTE.xlsx'), deixando o arquivo de verdade intocado.")
    parser.add_argument("--max-combinacao-pagamentos", dest="max_combinacao", type=int, default=None, help="Trava de segurança (opcional) pro tamanho máximo de combinação de pagamentos (2 ou mais) testada nas prioridades 3 e 4, quando uma OB isolada não bate com nenhum pagamento sozinho. Por padrão (não informado) não há trava: o escalonamento interno começa já no número de pagamentos pendentes daquela aba e só desce (até 6) se sobrarem OBs sem match. Informe um valor aqui só se quiser limitar isso por baixo, em abas com muitos pagamentos sem match, pra não deixar a busca cara demais.")

    args = parser.parse_args()

    caminho_original = Path(args.conciliacao)
    caminho_extratos = Path(args.extratos)

    if not caminho_original.exists():
        sys.exit(f"Planilha de conciliação não encontrada: {caminho_original}")
    if not caminho_extratos.exists():
        sys.exit(f"Planilha de extratos não encontrada: {caminho_extratos}")

    if args.copia:
        caminho_teste = caminho_original.with_name(
            caminho_original.stem + "_TESTE" + caminho_original.suffix
        )
        shutil.copyfile(caminho_original, caminho_teste)
        print(f"Trabalhando em cópia de teste: {caminho_teste}")
        print(f"(o arquivo original '{caminho_original}' NÃO é alterado)\n")
    else:
        caminho_teste = caminho_original
        print(f"Editando diretamente: {caminho_teste}\n")

    wb = openpyxl.load_workbook(caminho_teste, read_only=True)
    todas_abas = wb.sheetnames
    wb.close()

    abas_alvo = args.abas or list(todas_abas)

    for i, nome in enumerate(abas_alvo):
        if nome not in todas_abas:
            print(f"\nAba '{nome}' não encontrada em {caminho_teste}. Pulando.")
            continue

        try:
            rodar_aba(str(caminho_teste), str(caminho_extratos), nome, args.max_combinacao)
        except Exception:
            print(f"\n!!! Erro ao processar a aba '{nome}':")
            traceback.print_exc()
            if not args.sem_pausa:
                resp = input("\nContinuar para a próxima aba mesmo assim? [S/n] ").strip().lower()
                if resp == "n":
                    break
            continue

        ultima = (i == len(abas_alvo) - 1)
        if not args.sem_pausa and not ultima:
            resp = input(
                f"\nAba '{nome}' concluída. ENTER para ir para a próxima aba, "
                f"'q' para parar por aqui: "
            ).strip().lower()
            if resp == "q":
                break

    print(f"\nConcluído. Planilha resultante: {caminho_teste}")


if __name__ == "__main__":
    main()