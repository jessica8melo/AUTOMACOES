# -*- coding: utf-8 -*-
"""
Filtra o "Relatório de pagamento" pela coluna "Fornecedor", buscando as
linhas cujo Fornecedor corresponde a um determinado CAT (localidade) da
TABELA_LOCALIDADE (arquivo tabela_localidade.py).

Motivação:
    Na planilha de pagamento não existe uma coluna "CAT". O que existe é a
    coluna "Fornecedor", que às vezes contém o nome da localidade embutido
    em textos maiores, ex:
        "MUNICIPIO DE BAURU"
        "JOAO PESSOA SECRETARIA DE FINANCAS SEFIN"
    Este script recebe um CAT (ex: "Bauru"), confirma que ele existe na
    tabela de referência e retorna todas as linhas cujo Fornecedor contém
    esse nome de localidade (comparação sem distinção de maiúsculas/
    minúsculas e sem acentos).

Uso:
    python relatorio_pagamentos.py "Bauru" [caminho_do_arquivo.xlsx] [--sheet NOME_DA_ABA]

Se nenhum caminho de arquivo for informado, usa o arquivo padrão definido em
DEFAULT_PATH.

Requisito: o arquivo tabela_localidade.py deve estar na mesma pasta deste
script (ou no PYTHONPATH).
"""

import argparse
import unicodedata
from datetime import datetime, date

import openpyxl

from tabela_localidade import TABELA_LOCALIDADE, buscar_por_cat

DEFAULT_PATH = "Relatório de pagamento.xlsx"


def _normalizar(texto):
    """Remove acentos e converte para maiúsculas, para comparação tolerante."""
    if texto is None:
        return ""
    texto = str(texto).strip()
    sem_acento = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    return sem_acento.upper()


def _formatar_data(valor):
    if isinstance(valor, (datetime, date)):
        return valor.strftime("%d/%m/%Y")
    return valor


def resolver_cat(cat_informado):
    """
    Confirma que o CAT informado existe na TABELA_LOCALIDADE.
    A comparação ignora acentos e maiúsculas/minúsculas.
    Retorna o registro da tabela (dict) ou None se não encontrado.
    """
    alvo = _normalizar(cat_informado)
    for item in TABELA_LOCALIDADE:
        if item.get("cat") and _normalizar(item["cat"]) == alvo:
            return item
    return None


def encontrar_header(ws, colunas_esperadas=("Fornecedor",)):
    """Localiza a linha de cabeçalho e o índice (1-based) de cada coluna esperada."""
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30)):
        valores = {str(c.value).strip(): c.column for c in row if c.value is not None}
        if all(col in valores for col in colunas_esperadas):
            return row[0].row, valores
    raise ValueError(
        f"Não foi possível localizar uma linha de cabeçalho contendo as colunas: {colunas_esperadas}"
    )


def filtrar_por_cat(caminho_arquivo, cat_informado, nome_aba=None):
    registro_cat = resolver_cat(cat_informado)
    if registro_cat is None:
        raise ValueError(
            f'CAT "{cat_informado}" não foi encontrado na tabela_localidade.'
        )

    nome_cat_normalizado = _normalizar(registro_cat["cat"])

    wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
    ws = wb[nome_aba] if nome_aba else wb.active

    linha_header, col_idx = encontrar_header(ws)
    col_grupo = col_idx.get("Grupo Pagamentos")
    col_forn = col_idx["Fornecedor"]
    col_nff = col_idx.get("NFF")
    col_data_pgto = col_idx.get("Data Pagamento")
    col_valor = col_idx.get("Valor")

    resultados = []
    for row in ws.iter_rows(min_row=linha_header + 1, max_row=ws.max_row):
        fornecedor = row[col_forn - 1].value
        if fornecedor is None:
            continue
        if nome_cat_normalizado in _normalizar(fornecedor):
            resultados.append(
                {
                    "Grupo Pagamentos": row[col_grupo - 1].value if col_grupo else None,
                    "Fornecedor": fornecedor,
                    "NFF": row[col_nff - 1].value if col_nff else None,
                    "Data Pagamento": _formatar_data(row[col_data_pgto - 1].value) if col_data_pgto else None,
                    "Valor": row[col_valor - 1].value if col_valor else None,
                }
            )

    return registro_cat, resultados


def agrupar_valor_por_data(resultados):
    """
    Agrupa os resultados por "Data Pagamento" e soma a coluna "Valor" de cada grupo.
    Linhas sem "Data Pagamento" são agrupadas sob a chave None.
    Valores não numéricos são ignorados na soma.

    Retorna uma lista de dicts [{"Data Pagamento": ..., "Total": ...}, ...]
    ordenada por data (linhas sem data vão para o final).
    """
    somas = {}
    for item in resultados:
        data = item["Data Pagamento"]
        valor = item["Valor"]
        if isinstance(valor, (int, float)):
            somas[data] = somas.get(data, 0) + valor
        else:
            somas.setdefault(data, somas.get(data, 0))

    def chave_ordenacao(data):
        if data is None:
            return (1, "")
        # já vem formatada como "dd/mm/aaaa"; convertemos para ordenar corretamente
        try:
            dia, mes, ano = data.split("/")
            return (0, f"{ano}{mes}{dia}")
        except (AttributeError, ValueError):
            return (0, str(data))

    return [
        {"Data Pagamento": data, "Total": total}
        for data, total in sorted(somas.items(), key=lambda kv: chave_ordenacao(kv[0]))
    ]


def main():
    parser = argparse.ArgumentParser(
        description="Filtra a coluna Fornecedor do relatório de pagamento a partir de um CAT."
    )
    parser.add_argument("cat", help='Nome do CAT a filtrar (ex: "Bauru")')
    parser.add_argument("arquivo", nargs="?", default=DEFAULT_PATH, help="Caminho do arquivo .xlsx")
    parser.add_argument("--sheet", default=None, help="Nome da aba (padrão: aba ativa)")
    args = parser.parse_args()

    registro_cat, resultados = filtrar_por_cat(args.arquivo, args.cat, args.sheet)

    print(f'CAT: {registro_cat["cat"]}  |  Sigla: {registro_cat["sigla"]}  |  ORG: {registro_cat["org"]}\n')

    if not resultados:
        print("Nenhuma linha de Fornecedor encontrada para esse CAT.")
        return

    print(f"Total de ocorrências encontradas: {len(resultados)}\n")
    for i, item in enumerate(resultados, start=1):
        print(
            f"{i}. Fornecedor: {item['Fornecedor']} | Grupo: {item['Grupo Pagamentos']} | "
            f"NFF: {item['NFF']} | Data Pagamento: {item['Data Pagamento']} | Valor: {item['Valor']}"
        )

    print("\nSoma de Valor por Data Pagamento:\n")
    totais = agrupar_valor_por_data(resultados)
    for grupo in totais:
        data = grupo["Data Pagamento"] or "(sem data)"
        print(f"{data}: {grupo['Total']:.2f}")


if __name__ == "__main__":
    main()