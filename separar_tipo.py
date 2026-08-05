"""
Script para separar a planilha de conciliação em Recebimentos e Pagamentos,
mantendo tudo em UMA ÚNICA aba (Recebimentos primeiro, depois Pagamentos),
com uma borda mais grossa marcando a divisão entre os dois blocos.

Regra usada: qualquer valor da coluna "Tipo" que COMEÇA com "Recebimento"
vai para o bloco de Recebimentos; qualquer valor que COMEÇA com "Pagamento"
vai para o bloco de Pagamentos. Isso cobre variações como "Recebimento" e
"Recebimento Div." / "Pagamento" e "Pagamento Div.".

Uso:
    python separar_tipo.py caminho_da_planilha.xlsx
"""

import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FONT_NAME = "Arial"
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
# Cabeçalho de cada bloco (linha "Recebimentos" / "Pagamentos")
BLOCO_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
BLOCO_FILL_RECEB = PatternFill("solid", fgColor="2E7D32")
BLOCO_FILL_PAG = PatternFill("solid", fgColor="B71C1C")
NORMAL_FONT = Font(name=FONT_NAME, size=10)

THIN = Side(style="thin", color="B7B7B7")
THICK = Side(style="thick", color="000000")

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

NUM_COLS = 7  # A:G


def classificar(tipo: str) -> str:
    """Retorna 'Recebimento', 'Pagamento' ou 'Outro' com base no texto do Tipo."""
    if not tipo:
        return "Outro"
    t = str(tipo).strip().lower()
    if t.startswith("recebimento"):
        return "Recebimento"
    if t.startswith("pagamento"):
        return "Pagamento"
    return "Outro"


def localizar_bloco_ob(ws_origem):
    """Procura, na linha de cabeçalho, colunas chamadas 'OB' e 'VALOR' (nessa
    ordem, lado a lado) e devolve (headers, linhas) desse bloco, ou (None, None)
    se não encontrar."""
    header_row = ws_origem[1]
    col_ob = None
    for cell in header_row:
        if cell.value and str(cell.value).strip().upper() == "OB":
            col_ob = cell.column
            break
    if col_ob is None:
        return None, None

    col_valor = col_ob + 1
    valor_titulo = ws_origem.cell(row=1, column=col_valor).value
    if not valor_titulo or str(valor_titulo).strip().upper() != "VALOR":
        return None, None

    headers = [ws_origem.cell(row=1, column=col_ob).value,
               ws_origem.cell(row=1, column=col_valor).value]

    linhas = []
    for r in range(2, ws_origem.max_row + 1):
        v_ob = ws_origem.cell(row=r, column=col_ob).value
        v_valor = ws_origem.cell(row=r, column=col_valor).value
        if v_ob is None and v_valor is None:
            continue
        linhas.append((v_ob, v_valor))

    return headers, linhas


def escrever_cabecalho_colunas(ws, row, headers):
    for col, titulo in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=titulo)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def escrever_titulo_bloco(ws, row, texto, fill, num_cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    c = ws.cell(row=row, column=1, value=texto)
    c.font = BLOCO_FONT
    c.fill = fill
    c.alignment = CENTER
    for col in range(1, num_cols + 1):
        ws.cell(row=row, column=col).fill = fill


def escrever_linhas(ws, start_row, linhas, col_valor_idx=None):
    """Escreve as linhas de dados a partir de start_row. Retorna a última linha usada."""
    r = start_row
    for linha in linhas:
        for c_idx, valor in enumerate(linha, start=1):
            cell = ws.cell(row=r, column=c_idx, value=valor)
            cell.font = NORMAL_FONT
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            if col_valor_idx is not None and c_idx == col_valor_idx:
                cell.number_format = "#,##0.00"
                cell.alignment = RIGHT
            elif c_idx in (1, 4):  # Linha / Data Transação centralizadas
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT
        r += 1
    return r - 1


def aplicar_borda_grossa_inferior(ws, row, num_cols):
    """Engrossa a borda inferior de toda a linha indicada, marcando a divisão
    entre o bloco de Recebimentos e o bloco de Pagamentos."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        existing = cell.border
        cell.border = Border(
            left=existing.left, right=existing.right,
            top=existing.top, bottom=THICK,
        )


def main(caminho):
    wb = openpyxl.load_workbook(caminho)
    ws_origem = wb["Conciliação"] if "Conciliação" in wb.sheetnames else wb.active

    headers = [c.value for c in ws_origem[1][:NUM_COLS]]  # A:G

    recebimentos = []
    pagamentos = []
    outros = []

    for row in ws_origem.iter_rows(min_row=2, max_col=NUM_COLS, values_only=True):
        if row[0] is None:
            continue
        tipo = row[1]
        categoria = classificar(tipo)
        if categoria == "Recebimento":
            recebimentos.append(row)
        elif categoria == "Pagamento":
            pagamentos.append(row)
        else:
            outros.append(row)

    headers_ob, linhas_ob = localizar_bloco_ob(ws_origem)

    nome_aba = "Recebimentos e Pagamentos"
    if nome_aba in wb.sheetnames:
        del wb[nome_aba]
    ws = wb.create_sheet(nome_aba)

    linha_atual = 1

    # --- Bloco Recebimentos ---
    escrever_titulo_bloco(ws, linha_atual, f"RECEBIMENTOS ({len(recebimentos)})",
                           BLOCO_FILL_RECEB, NUM_COLS)
    linha_atual += 1
    escrever_cabecalho_colunas(ws, linha_atual, headers)
    linha_header_receb = linha_atual
    linha_atual += 1
    ultima_linha_receb = escrever_linhas(ws, linha_atual, recebimentos, col_valor_idx=5)
    if not recebimentos:
        ultima_linha_receb = linha_atual - 1

    # Borda grossa separando os dois blocos
    linha_divisoria = max(ultima_linha_receb, linha_header_receb)
    aplicar_borda_grossa_inferior(ws, linha_divisoria, NUM_COLS)

    linha_atual = linha_divisoria + 2  # pula uma linha em branco

    # --- Bloco Pagamentos ---
    linha_titulo_pag = linha_atual
    escrever_titulo_bloco(ws, linha_atual, f"PAGAMENTOS ({len(pagamentos)})",
                           BLOCO_FILL_PAG, NUM_COLS)
    linha_atual += 1
    escrever_cabecalho_colunas(ws, linha_atual, headers)
    linha_atual += 1
    inicio_dados_pag = linha_atual
    ultima_linha_pag = escrever_linhas(ws, linha_atual, pagamentos, col_valor_idx=5)

    # Bloco extra OB/VALOR ao lado do bloco de Pagamentos (colunas I:J)
    if headers_ob is not None:
        col_inicial_ob = NUM_COLS + 2  # coluna I
        escrever_cabecalho_colunas_extra = None
        # título do mini-bloco, alinhado com o título de Pagamentos
        ws.merge_cells(start_row=linha_titulo_pag, start_column=col_inicial_ob,
                        end_row=linha_titulo_pag, end_column=col_inicial_ob + 1)
        c = ws.cell(row=linha_titulo_pag, column=col_inicial_ob, value="OB / VALOR")
        c.font = BLOCO_FONT
        c.fill = BLOCO_FILL_PAG
        c.alignment = CENTER
        ws.cell(row=linha_titulo_pag, column=col_inicial_ob + 1).fill = BLOCO_FILL_PAG

        for col_off, titulo in enumerate(headers_ob):
            cc = ws.cell(row=linha_titulo_pag + 1, column=col_inicial_ob + col_off, value=titulo)
            cc.font = HEADER_FONT
            cc.fill = HEADER_FILL
            cc.alignment = CENTER
            cc.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

        for i, (v_ob, v_valor) in enumerate(linhas_ob):
            r = inicio_dados_pag + i
            c_ob = ws.cell(row=r, column=col_inicial_ob, value=v_ob)
            c_ob.font = NORMAL_FONT
            c_ob.alignment = CENTER
            c_ob.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

            c_val = ws.cell(row=r, column=col_inicial_ob + 1, value=v_valor)
            c_val.font = NORMAL_FONT
            c_val.number_format = "#,##0.00"
            c_val.alignment = RIGHT
            c_val.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

        letra_inicial = ws.cell(row=1, column=col_inicial_ob).column_letter
        ws.column_dimensions[letra_inicial].width = 14
        ws.column_dimensions[ws.cell(row=1, column=col_inicial_ob + 1).column_letter].width = 14

    linha_final_pag = ultima_linha_pag

    # --- Bloco Outros (se houver) ---
    if outros:
        linha_atual = linha_final_pag + 2
        escrever_titulo_bloco(ws, linha_atual, f"OUTROS ({len(outros)})",
                               PatternFill("solid", fgColor="616161"), NUM_COLS)
        linha_atual += 1
        escrever_cabecalho_colunas(ws, linha_atual, headers)
        linha_atual += 1
        escrever_linhas(ws, linha_atual, outros, col_valor_idx=5)

    larguras = {"A": 7, "B": 18, "C": 11, "D": 15, "E": 13, "F": 18, "G": 32}
    for col, w in larguras.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    wb.save(caminho)

    print(f"Recebimentos: {len(recebimentos)} linhas")
    print(f"Pagamentos:   {len(pagamentos)} linhas")
    if outros:
        print(f"Outros:       {len(outros)} linhas (não classificados)")
    if headers_ob is not None:
        print(f"Bloco OB/VALOR: {len(linhas_ob)} linhas ao lado do bloco de Pagamentos")
    else:
        print("Bloco OB/VALOR: não encontrado na planilha de origem (colunas 'OB'/'VALOR' não localizadas)")
    print(f"Tudo reunido na aba '{nome_aba}'.")


if __name__ == "__main__":
    caminho = sys.argv[1] if len(sys.argv) > 1 else "conciliacao.xlsx"
    main(caminho)