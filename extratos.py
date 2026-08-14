"""
Passo 3 (opcional, roda só se --extratos for informado): cruza TODOS os
pagamentos da aba (tenham ou não sido conciliados com uma OB no Passo 2)
com a tabela de PAGAMENTOS da planilha de extratos (saida.xlsx, gerada por
outro script a partir dos extratos bancários), usando como chave a
combinação Data Transação + Quantia (comparado com Data + Valor (R$) do
extrato). Para cada pagamento:

    1) Se achar uma linha do extrato com a mesma Data+Valor e o Histórico
       dela começar com "Tarifa" ou "Tar agrupad" (ex.: "Tar agrupadas"),
       preenche a coluna "Status" com "TARIFA" (independentemente do
       pagamento ter sido conciliado com uma OB). Ver `_eh_historico_tarifa`.
    2) Se achar e o Histórico for uma transferência interna entre contas do
       grupo BBTS (ver `_eh_transferencia_entre_contas_bbts`), preenche
       "Status" no formato "Conta_origem / Conta_destino" (também
       independentemente de conciliação com OB, igual TARIFA).
    3) Se achar e o Histórico for exatamente "Pagamento" ou "Pagamentos
       Diversos" (sem contar maiúsculas/minúsculas e espaços), a linha é
       ignorada - nada é preenchido.
    4) Se achar e não for nenhum dos casos acima, preenche "Status" com o
       texto do Histórico do extrato. A coluna "Justificativas - Não
       reconciliadas" só recebe uma justificativa quando o pagamento NÃO
       tiver sido conciliado com nenhuma OB no Passo 2 (pra pagamentos já
       conciliados, só o Status é preenchido).
    5) Se não achar nenhuma linha do extrato com aquela Data+Valor, a linha
       é deixada como está (fica sinalizada no log para conferência
       manual).

Se mais de uma linha do extrato bater com a mesma Data+Valor (ambíguo), a
primeira encontrada é usada, mas o caso é destacado no log para revisão.

Além disso, para a tabela de RECEBIMENTOS da mesma aba (quando existir),
cada linha também é cruzada por Data+Valor com a tabela de PAGAMENTOS do
extrato. Quando o Histórico encontrado for uma transferência interna entre
contas do grupo BBTS (mesmo critério do item 2 acima - ver
`_eh_transferencia_entre_contas_bbts`), a coluna "Status" é preenchida no
formato "Conta_origem / Conta_destino", usando o nome da aba do extrato de
onde saiu o dinheiro (conta origem) e o número extraído do campo Documento
dessa linha (conta destino) - ver `atualizar_status_recebimentos_aba`.
Outras correspondências no lado de Recebimentos (ex.: Pix) são ignoradas.

Uma transferência é considerada interna (entre contas do grupo BBTS) quando
o Histórico começa com "Transferência enviada" E (contém "BB TEC" OU o
número de conta extraído do campo Documento bate com uma das contas
presentes nas abas do saida.xlsx) - isso cobre tanto o caso mais comum
("... BB TEC ...") quanto variações de texto que não mencionam "BB TEC"
mas cujo Documento aponta pra uma conta BBTS conhecida (ex.: "Transferência
enviada 31/07 09:46 00042318949000184", sem "BB TEC" no texto).
"""

import datetime
import re

import openpyxl

from processar_pagamentos import BOLD_FONT, localizar_titulo, _centavos, _normalizar_abas
from conciliacao import conciliar_pagamentos_obs_aba

MESES_PT = {
    "jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
    "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12,
}

COL_STATUS = 6
COL_JUSTIFICATIVA = 7

HISTORICOS_IGNORADOS = {"pagamento", "pagamentos diversos"}

# Padrões de Histórico do extrato que devem ser tratados como tarifa
# bancária (Status = "TARIFA"). Não são checados só no INÍCIO do texto,
# porque alguns bancos embutem o trecho no meio de uma descrição maior,
# ex.: "Débito Serviço Cobrança Tar. agrupadas - ocorrencia 30/07/2026".
# Cobre tanto "Tarifa..." quanto "Tar agrupadas"/"Tar. agrupadas" (com ou
# sem o ponto depois de "Tar").
_RE_TARIFA = re.compile(r"\btarifa\b", re.IGNORECASE)
_RE_TAR_AGRUPADAS = re.compile(r"\btar\.?\s*agrupad", re.IGNORECASE)


def _eh_historico_tarifa(historico_str):
    """True se o histórico do extrato deve ser tratado como tarifa
    bancária (Status = 'TARIFA'), buscando os padrões em qualquer parte do
    texto (não só no início)."""
    return bool(_RE_TARIFA.search(historico_str) or _RE_TAR_AGRUPADAS.search(historico_str))


# Histórico de uma transferência interna entre contas do grupo BBTS, ex.:
# "Transferência enviada 27/07 11:59 BB TEC 00042318949000184". Tem que
# COMEÇAR com "Transferência enviada" (pra não pegar "Transferência
# recebida", que é o espelho do lado de quem recebeu). Consideramos
# transferência interna quando o texto contém "BB TEC" OU quando o número
# de conta extraído do campo Documento bate com uma conta conhecida
# (presente nas abas do saida.xlsx) - esse segundo critério cobre casos
# como "Transferência enviada 31/07 09:47 BB T SERVICOS S.A." ou
# "Transferência enviada 31/07 09:46 00042318949000184", que não têm "BB
# TEC" no texto mas cujo Documento aponta pra uma conta BBTS conhecida.
# Transferências pra terceiros (ex.: "MICHEL ANDRADE HAMU") continuam de
# fora, porque o Documento delas não bate com nenhuma conta conhecida.
_RE_TRANSFERENCIA_ENVIADA = re.compile(r"^transfer[êe]ncia\s+enviada\b", re.IGNORECASE)
_RE_BB_TEC = re.compile(r"\bbb\s*tec\b", re.IGNORECASE)

# Dígitos do campo Documento de uma linha de transferência, ex.:
# "553.309.000.205.826" -> os últimos 6 dígitos ("205826") são o número da
# conta de destino (sem o dígito verificador).
_RE_DIGITOS = re.compile(r"\d+")

# Nome de aba do saida.xlsx = número da conta, ex.: "205826-X".
_RE_NOME_CONTA = re.compile(r"^(\d{6})-(\w+)$")


def _eh_transferencia_entre_contas_bbts(historico_str, documento, mapa_contas):
    """True se o histórico do extrato for uma 'Transferência enviada...'
    que corresponda a uma transferência interna entre contas do grupo
    BBTS: ou o texto contém 'BB TEC', ou o número de conta extraído do
    campo Documento (ver `_conta_destino_do_documento`) bate com uma conta
    conhecida em `mapa_contas`."""
    if not _RE_TRANSFERENCIA_ENVIADA.match(historico_str):
        return False
    if _RE_BB_TEC.search(historico_str):
        return True
    digitos = "".join(_RE_DIGITOS.findall(str(documento or "")))
    return len(digitos) >= 6 and digitos[-6:] in mapa_contas


def _mapear_contas(nomes_abas_extrato):
    """Monta um mapa {6 dígitos -> 'NNNNNN-D'} a partir dos nomes das abas
    do saida.xlsx (cada aba é nomeada com o número da conta, incluindo o
    dígito verificador, ex.: '205826-X'). Usado pra recuperar o dígito
    verificador da conta de destino extraída do campo Documento."""
    mapa = {}
    for nome in nomes_abas_extrato:
        m = _RE_NOME_CONTA.match(nome.strip())
        if m:
            mapa[m.group(1)] = nome
    return mapa


def _conta_destino_do_documento(documento, mapa_contas):
    """A partir do campo Documento de uma linha de 'Transferência enviada'
    (ex.: '553.309.000.205.826'), extrai os últimos 6 dígitos (número da
    conta de destino, sem dígito verificador) e tenta casar com uma conta
    conhecida (`mapa_contas`, ver `_mapear_contas`) pra devolver o número
    completo, com dígito verificador (ex.: '205826-X'). Se o Documento não
    tiver dígitos suficientes, devolve None; se tiver mas não achar a conta
    no mapa, devolve só os 6 dígitos (sem dígito verificador)."""
    digitos = "".join(_RE_DIGITOS.findall(str(documento or "")))
    if len(digitos) < 6:
        return None
    ultimos6 = digitos[-6:]
    return mapa_contas.get(ultimos6, ultimos6)

# Abas da planilha de conciliação seguem o padrão "dd.mm.aa" (ex.: "07.08.25"),
# que é a data do dia a que aquela aba se refere.
_ABA_DATA_RE = re.compile(r"^(\d{1,2})\.(\d{1,2})\.(\d{2,4})$")


def _data_aba_para_tupla(nome_aba):
    """Tenta interpretar o NOME da aba (ex.: '07.08.25') como uma data no
    formato 'dd.mm.aa' ou 'dd.mm.aaaa'. Devolve (ano, mes, dia), ou None se
    o nome da aba não seguir esse padrão (ex.: 'Modelo DICOS', 'Planilha24')
    ou não for uma data válida (ex.: '32.13.99')."""
    m = _ABA_DATA_RE.match(nome_aba.strip())
    if not m:
        return None
    dia, mes, ano = (int(x) for x in m.groups())
    if ano < 100:
        ano += 2000
    try:
        datetime.date(ano, mes, dia)
    except ValueError:
        return None
    return (ano, mes, dia)


def _data_para_tupla(valor):
    """Aceita 'dd/mmm/aa' (ex 27/jul/26), 'dd/mm/aaaa' (ex 27/07/2026) e
    também células de data reais do Excel (datetime.datetime / datetime.date,
    como vem em algumas abas de conciliacao.xlsx). Devolve (ano, mes, dia),
    ou None se não conseguir interpretar."""
    if valor is None:
        return None

    if isinstance(valor, datetime.datetime):
        return (valor.year, valor.month, valor.day)
    if isinstance(valor, datetime.date):
        return (valor.year, valor.month, valor.day)

    partes = str(valor).strip().split("/")
    if len(partes) != 3:
        return None
    dia_str, mes_str, ano_str = partes
    try:
        dia = int(dia_str)
    except ValueError:
        return None

    if mes_str.isdigit():
        mes = int(mes_str)
    else:
        mes = MESES_PT.get(mes_str.strip().lower()[:3])
        if mes is None:
            return None

    try:
        ano = int(ano_str)
    except ValueError:
        return None
    if ano < 100:
        ano += 2000

    return (ano, mes, dia)


def _ler_pagamentos_extrato_aba(ws):
    """Lê a tabela de PAGAMENTOS de uma aba de saida.xlsx (colunas
    Data/Histórico/Documento/Valor). Devolve [(data, histórico, documento,
    valor)]."""
    linha_titulo = localizar_titulo(ws, "PAGAMENTOS")
    if linha_titulo is None:
        return []

    linhas = []
    r = linha_titulo + 2  # +1 cabeçalho de colunas, +2 primeira linha de dado
    while True:
        data = ws.cell(row=r, column=1).value
        if data is None:
            break
        historico = ws.cell(row=r, column=2).value
        documento = ws.cell(row=r, column=3).value
        valor = ws.cell(row=r, column=4).value
        if isinstance(valor, (int, float)):
            linhas.append((data, historico, documento, valor))
        r += 1
    return linhas


def _indexar_extrato(caminho_extratos):
    """Monta um índice (ano, mes, dia, valor_em_centavos) -> [(aba,
    histórico, documento), ...] a partir de todas as abas de PAGAMENTOS em
    saida.xlsx, e também o mapa de contas conhecidas (ver `_mapear_contas`).
    Devolve (indice, mapa_contas)."""
    wb = openpyxl.load_workbook(caminho_extratos, data_only=True)
    indice = {}
    for nome in wb.sheetnames:
        ws = wb[nome]
        for data, historico, documento, valor in _ler_pagamentos_extrato_aba(ws):
            chave_data = _data_para_tupla(data)
            if chave_data is None:
                continue
            chave = (*chave_data, _centavos(valor))
            indice.setdefault(chave, []).append((nome, historico, documento))
    return indice, _mapear_contas(wb.sheetnames)


def atualizar_status_sem_conciliacao_aba(ws, pagamentos_todos, linhas_conciliadas, indice_extrato, mapa_contas):
    """Para CADA pagamento da aba (linha, quantia) - conciliado com OB ou
    não - procura a mesma Data+Valor na tabela de extratos:
      - se achar e o Histórico começar com 'Tarifa' ou 'Tar agrupad'
        (ver `_eh_historico_tarifa`) -> Status = 'TARIFA' (independe de ter
        sido conciliado com OB).
      - se achar e o Histórico for uma transferência interna entre contas
        do grupo BBTS (ver `_eh_transferencia_entre_contas_bbts`) -> Status
        = 'Conta_origem / Conta_destino' (independe de ter sido conciliado
        com OB, igual TARIFA).
      - se achar e o Histórico for exatamente 'Pagamento' ou 'Pagamentos
        Diversos' -> nada é alterado (ignorado).
      - se achar e não for nenhum dos casos acima -> Status = texto do
        Histórico do extrato; a Justificativa só é preenchida se esse
        pagamento NÃO estiver em `linhas_conciliadas` (ou seja, se ele
        tiver ficado sem OB).
      - se não achar nenhuma correspondência -> nada é alterado (fica para
        conferência manual).
    Devolve um relatório (listas de linhas tratadas como tarifa, como
    transferência interna, mandadas pra contas a pagar, com status
    preenchido mas já conciliadas com OB, ignoradas por serem histórico
    genérico, sem correspondência no extrato e ambíguas)."""
    tratados_tarifa = []
    tratados_transferencia = []
    tratados_contas_a_pagar = []
    tratados_conciliados_com_status = []
    ignorados_historico_generico = []
    sem_correspondencia = []
    ambiguos = []

    for linha, quantia in pagamentos_todos:
        data_transacao = ws.cell(row=linha, column=4).value
        chave_data = _data_para_tupla(data_transacao)
        if chave_data is None:
            sem_correspondencia.append((linha, quantia))
            continue

        chave = (*chave_data, _centavos(quantia))
        candidatos = indice_extrato.get(chave)
        if not candidatos:
            sem_correspondencia.append((linha, quantia))
            continue

        if len(candidatos) > 1:
            ambiguos.append((linha, quantia, candidatos))

        aba_origem, historico, documento = candidatos[0]
        historico_str = str(historico or "").strip()

        if _eh_historico_tarifa(historico_str):
            cel_status = ws.cell(row=linha, column=COL_STATUS, value="TARIFA")
            cel_status.font = BOLD_FONT
            tratados_tarifa.append((linha, quantia, aba_origem, historico_str))
            continue

        if _eh_transferencia_entre_contas_bbts(historico_str, documento, mapa_contas):
            conta_destino = _conta_destino_do_documento(documento, mapa_contas)
            if conta_destino is not None:
                status = f"{aba_origem} / {conta_destino}"
                cel_status = ws.cell(row=linha, column=COL_STATUS, value=status)
                cel_status.font = BOLD_FONT
                tratados_transferencia.append((linha, quantia, aba_origem, conta_destino))
                continue
            # Documento sem dígitos suficientes: cai no tratamento genérico
            # abaixo (Status = texto do Histórico), pra não perder o registro.

        if historico_str.lower() in HISTORICOS_IGNORADOS:
            ignorados_historico_generico.append((linha, quantia, aba_origem, historico_str))
            continue

        ws.cell(row=linha, column=COL_STATUS, value=historico_str)
        if linha in linhas_conciliadas:
            tratados_conciliados_com_status.append((linha, quantia, aba_origem, historico_str))
        else:
            tratados_contas_a_pagar.append((linha, quantia, aba_origem, historico_str))

    return {
        "tarifa": tratados_tarifa,
        "transferencia": tratados_transferencia,
        "contas_a_pagar": tratados_contas_a_pagar,
        "conciliados_com_status": tratados_conciliados_com_status,
        "ignorados_historico_generico": ignorados_historico_generico,
        "sem_correspondencia": sem_correspondencia,
        "ambiguos": ambiguos,
    }


def _ler_recebimentos_aba(ws):
    """Lê a tabela de RECEBIMENTOS da aba de conciliação (mesmo layout de
    colunas A:G da tabela de Pagamentos - ver `estilos.NUM_COLS_PRINCIPAL`).
    Devolve [(linha, quantia)], igual a `conciliacao._linhas_pagamentos`."""
    linha_titulo = localizar_titulo(ws, "RECEBIMENTOS")
    if linha_titulo is None:
        return []

    linhas = []
    r = linha_titulo + 2  # +1 = cabeçalho de colunas, +2 = primeira linha de dado
    while ws.cell(row=r, column=1).value is not None:
        quantia = ws.cell(row=r, column=5).value  # coluna E = Quantia
        if quantia is not None:
            linhas.append((r, quantia))
        r += 1
    return linhas


def atualizar_status_recebimentos_aba(ws, indice_extrato, mapa_contas):
    """Para cada linha da tabela de RECEBIMENTOS da aba, procura a mesma
    Data+Valor na tabela de PAGAMENTOS do extrato (mesmo índice usado para
    os Pagamentos - Passo 3 de `atualizar_status_sem_conciliacao_aba`).

    Só age quando a correspondência encontrada for uma transferência
    interna entre contas do grupo BBTS (ver
    `_eh_transferencia_entre_contas_bbts` - Histórico começando com
    'Transferência enviada' e contendo 'BB TEC' OU cujo Documento aponte
    pra uma conta conhecida); outras correspondências (Pix, etc.) são
    deixadas como estão. Quando aplicável, preenche a coluna 'Status' com
    'Conta_origem / Conta_destino':
      - Conta_origem = nome da aba do extrato de onde saiu a transferência
        (a própria aba já é nomeada com o número da conta, ex.: '200000-8');
      - Conta_destino = número extraído do campo Documento dessa linha do
        extrato, com o dígito verificador recuperado via `mapa_contas`
        quando possível (ver `_conta_destino_do_documento`).

    A coluna Justificativas não é alterada.

    Devolve um relatório com as linhas marcadas, as ambíguas (mais de uma
    linha do extrato de transferência BBTS batendo com a mesma Data+Valor)
    e as que bateram mas cujo Documento não deu pra interpretar."""
    tratados = []
    ambiguos = []
    documento_invalido = []

    for linha, quantia in _ler_recebimentos_aba(ws):
        data_transacao = ws.cell(row=linha, column=4).value
        chave_data = _data_para_tupla(data_transacao)
        if chave_data is None:
            continue

        chave = (*chave_data, _centavos(quantia))
        candidatos = indice_extrato.get(chave)
        if not candidatos:
            continue

        candidatos_transferencia = [
            (aba_origem, historico, documento)
            for aba_origem, historico, documento in candidatos
            if _eh_transferencia_entre_contas_bbts(str(historico or "").strip(), documento, mapa_contas)
        ]
        if not candidatos_transferencia:
            continue
        if len(candidatos_transferencia) > 1:
            ambiguos.append((linha, quantia, candidatos_transferencia))

        conta_origem, historico, documento = candidatos_transferencia[0]
        conta_destino = _conta_destino_do_documento(documento, mapa_contas)
        if conta_destino is None:
            documento_invalido.append((linha, quantia, conta_origem, documento))
            continue

        status = f"{conta_origem} / {conta_destino}"
        cel_status = ws.cell(row=linha, column=COL_STATUS, value=status)
        cel_status.font = BOLD_FONT
        tratados.append((linha, quantia, conta_origem, conta_destino))

    return {
        "transferencias": tratados,
        "ambiguos": ambiguos,
        "documento_invalido": documento_invalido,
    }


def atualizar_status_sem_conciliacao(caminho_conciliacao, caminho_extratos, max_combinacao=None, max_combinacao_obs=6, abas=None):
    """Roda a conciliação normal (pagamento <-> OB) em TODAS as abas alvo e,
    além disso, para cada aba cuja DATA (extraída do próprio nome da aba,
    ex.: '07.08.25') tiver pelo menos um lançamento na mesma data dentro da
    planilha de extratos, cruza os pagamentos dessa aba com o extrato por
    Data+Valor para preencher Status/Justificativas (Etapa 3).

    Abas cuja data não bate com nenhuma data presente no extrato (ou cujo
    nome nem sequer parece uma data) têm a Etapa 3 pulada - a conciliação
    normal (Passo 2) continua rodando normalmente nelas. Salva o resultado
    em `caminho_conciliacao`."""
    indice_extrato, mapa_contas = _indexar_extrato(caminho_extratos)
    datas_no_extrato = {chave[:3] for chave in indice_extrato}

    wb = openpyxl.load_workbook(caminho_conciliacao)
    abas_selecionadas = _normalizar_abas(abas)
    abas_alvo = abas_selecionadas or wb.sheetnames

    for nome in abas_alvo:
        if nome not in wb.sheetnames:
            print(f"Aba '{nome}' não encontrada em {caminho_conciliacao}. Ignorando.")
            continue

        ws = wb[nome]
        resultado_conciliacao = conciliar_pagamentos_obs_aba(
            ws, max_combinacao=max_combinacao, max_combinacao_obs=max_combinacao_obs
        )
        if resultado_conciliacao is None:
            continue

        n_obs_conciliadas = sum(len(grupo["ob_rows"]) for grupo in resultado_conciliacao["grupos"])
        print(f"\n[{nome}] OBs conciliadas: {n_obs_conciliadas} de {resultado_conciliacao['total_obs']} (em {len(resultado_conciliacao['grupos'])} grupo(s))")

        # --- Etapa 3 só roda se a data da ABA existir no extrato ---
        data_aba = _data_aba_para_tupla(nome)
        if data_aba is None:
            print(f"  ⏭️  Etapa 3 pulada: nome da aba '{nome}' não parece uma data (formato esperado dd.mm.aa).")
            continue
        if data_aba not in datas_no_extrato:
            ano, mes, dia = data_aba
            print(f"  ⏭️  Etapa 3 pulada: nenhum lançamento do extrato encontrado para {dia:02d}/{mes:02d}/{ano}.")
            continue

        linhas_conciliadas = {
            linha
            for grupo in resultado_conciliacao["grupos"]
            for linha in grupo["pagamento_rows"]
        }

        relatorio = atualizar_status_sem_conciliacao_aba(
            ws, resultado_conciliacao["pagamentos_todos"], linhas_conciliadas, indice_extrato, mapa_contas
        )

        print(f"\n[{nome}] Total de pagamentos na aba: {len(resultado_conciliacao['pagamentos_todos'])} "
              f"(sem OB: {len(resultado_conciliacao['pagamentos_sem_ob'])})")
        print(f"  -> Marcados como TARIFA: {len(relatorio['tarifa'])}")
        for linha, quantia, aba_origem, historico in relatorio["tarifa"]:
            print(f"     linha {linha} (R$ {quantia:.2f}) <- [{aba_origem}] {historico}")

        print(f"  -> Marcados como transferência interna BBTS: {len(relatorio['transferencia'])}")
        for linha, quantia, conta_origem, conta_destino in relatorio["transferencia"]:
            print(f"     linha {linha} (R$ {quantia:.2f}) <- {conta_origem} / {conta_destino}")

        print(f"  -> Marcados sem OB: {len(relatorio['contas_a_pagar'])}")
        for linha, quantia, aba_origem, historico in relatorio["contas_a_pagar"]:
            print(f"     linha {linha} (R$ {quantia:.2f}) <- [{aba_origem}] {historico}")

        print(f"  -> Status preenchido (já conciliado com OB, sem mexer na Justificativa): {len(relatorio['conciliados_com_status'])}")
        for linha, quantia, aba_origem, historico in relatorio["conciliados_com_status"]:
            print(f"     linha {linha} (R$ {quantia:.2f}) <- [{aba_origem}] {historico}")

        if relatorio["ignorados_historico_generico"]:
            print(f"  -> Ignorados (Histórico = Pagamento/Pagamentos Diversos): {len(relatorio['ignorados_historico_generico'])}")
            for linha, quantia, aba_origem, historico in relatorio["ignorados_historico_generico"]:
                print(f"     linha {linha} (R$ {quantia:.2f}) <- [{aba_origem}] {historico}")

        if relatorio["sem_correspondencia"]:
            print(f"  ⚠️  Sem correspondência nenhuma no extrato: {len(relatorio['sem_correspondencia'])}")
            for linha, quantia in relatorio["sem_correspondencia"]:
                print(f"     linha {linha} (R$ {quantia:.2f}) - conferir manualmente")

        if relatorio["ambiguos"]:
            print(f"  ⚠️  Data+Valor batendo com MAIS de uma linha do extrato (usada a primeira encontrada): {len(relatorio['ambiguos'])}")
            for linha, quantia, candidatos in relatorio["ambiguos"]:
                print(f"     linha {linha} (R$ {quantia:.2f}) candidatos: {candidatos}")

        # --- Recebimentos: só transferências internas BBTS (Status = Conta_origem / Conta_destino) ---
        relatorio_recebimentos = atualizar_status_recebimentos_aba(ws, indice_extrato, mapa_contas)

        if relatorio_recebimentos["transferencias"]:
            print(f"  -> Recebimentos marcados como transferência interna BBTS: {len(relatorio_recebimentos['transferencias'])}")
            for linha, quantia, conta_origem, conta_destino in relatorio_recebimentos["transferencias"]:
                print(f"     linha {linha} (R$ {quantia:.2f}) <- {conta_origem} / {conta_destino}")

        if relatorio_recebimentos["ambiguos"]:
            print(f"  ⚠️  Recebimento com MAIS de uma transferência BBTS batendo na mesma Data+Valor (usada a primeira): {len(relatorio_recebimentos['ambiguos'])}")
            for linha, quantia, candidatos in relatorio_recebimentos["ambiguos"]:
                print(f"     linha {linha} (R$ {quantia:.2f}) candidatos: {candidatos}")

        if relatorio_recebimentos["documento_invalido"]:
            print(f"  ⚠️  Recebimento batendo com transferência BBTS mas com Documento sem dígitos suficientes: {len(relatorio_recebimentos['documento_invalido'])}")
            for linha, quantia, conta_origem, documento in relatorio_recebimentos["documento_invalido"]:
                print(f"     linha {linha} (R$ {quantia:.2f}) origem {conta_origem}, Documento={documento!r} - conferir manualmente")

    wb.save(caminho_conciliacao)