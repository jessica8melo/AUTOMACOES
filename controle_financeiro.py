# -*- coding: utf-8 -*-
"""
Analisa a planilha de controle financeiro e retorna todas as linhas em que a
célula da coluna "OB" está com preenchimento AMARELO **e** o valor da célula
está vazio (sem número de OB preenchido).

Para cada ocorrência, retorna os valores das colunas:
    - Data
    - Descrição
    - Total OB

Uso:
    python controle_financeiro.py controle-financeiro-2026-2.xlsx [--sheet NOME_DA_ABA]

Se nenhum caminho for informado, usa o arquivo padrão definido em DEFAULT_PATH.
"""

import sys
import argparse
from datetime import datetime, date

import openpyxl

DEFAULT_PATH = "controle-financeiro-2026-2.xlsx"

# Cor de referência para "amarelo". O Excel/LibreOffice normalmente grava
# amarelo puro como FFFF00 (com prefixo de alpha "FF" -> "FFFFFF00").
AMARELO_RGB = "FFFF00"


def _rgb_da_celula(cell):
    """Extrai o RGB (6 dígitos hex, sem alpha) do preenchimento de uma célula."""
    fill = cell.fill
    if fill is None or fill.patternType is None:
        return None
    fg = fill.fgColor
    if fg is None or fg.type != "rgb" or not fg.rgb:
        return None
    rgb = fg.rgb
    # rgb geralmente vem como "AARRGGBB" (8 dígitos); pegamos só "RRGGBB"
    return rgb[-6:].upper() if rgb else None


def _e_amarelo(cell, tolerancia=30):
    """
    Verifica se a célula está pintada de amarelo.
    Aceita pequenas variações de tom (tolerância nos componentes RGB),
    já que planilhas reais às vezes usam amarelos ligeiramente diferentes.
    """
    rgb = _rgb_da_celula(cell)
    if rgb is None:
        return False
    try:
        r, g, b = int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)
    except ValueError:
        return False

    alvo_r, alvo_g, alvo_b = 255, 255, 0
    return (
        abs(r - alvo_r) <= tolerancia
        and abs(g - alvo_g) <= tolerancia
        and b <= tolerancia
    )


def _formatar_data(valor):
    if isinstance(valor, (datetime, date)):
        return valor.strftime("%d/%m/%Y")
    return valor


def encontrar_header(ws, colunas_esperadas=("OB", "Data", "Descrição", "Total OB")):
    """Localiza a linha de cabeçalho e o índice (1-based) de cada coluna esperada."""
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20)):
        valores = {str(c.value).strip(): c.column for c in row if c.value is not None}
        if all(col in valores for col in colunas_esperadas):
            linha_header = row[0].row
            return linha_header, valores
    raise ValueError(
        f"Não foi possível localizar uma linha de cabeçalho contendo as colunas: {colunas_esperadas}"
    )


def analisar(caminho_arquivo, nome_aba=None):
    wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
    ws = wb[nome_aba] if nome_aba else wb.active

    linha_header, col_idx = encontrar_header(ws)
    col_ob = col_idx["OB"]
    col_data = col_idx["Data"]
    col_desc = col_idx["Descrição"]
    col_total = col_idx["Total OB"]

    resultados = []
    for row in ws.iter_rows(min_row=linha_header + 1, max_row=ws.max_row):
        cell_ob = row[col_ob - 1]
        vazio = cell_ob.value is None or str(cell_ob.value).strip() == ""
        if vazio and _e_amarelo(cell_ob):
            resultados.append(
                {
                    "Data": _formatar_data(row[col_data - 1].value),
                    "Descrição": row[col_desc - 1].value,
                    "Total OB": row[col_total - 1].value,
                }
            )

    return resultados


def main():
    parser = argparse.ArgumentParser(description="Extrai ocorrências com OB pintada de amarelo.")
    parser.add_argument("arquivo", nargs="?", default=DEFAULT_PATH, help="Caminho do arquivo .xlsx")
    parser.add_argument("--sheet", default=None, help="Nome da aba (padrão: aba ativa)")
    args = parser.parse_args()

    resultados = analisar(args.arquivo, args.sheet)

    if not resultados:
        print("Nenhuma ocorrência com célula OB amarela foi encontrada.")
        return

    print(f"Total de ocorrências encontradas: {len(resultados)}\n")
    for i, item in enumerate(resultados, start=1):
        print(f"{i}. Data: {item['Data']} | Descrição: {item['Descrição']} | Total OB: {item['Total OB']}")


if __name__ == "__main__":
    main()