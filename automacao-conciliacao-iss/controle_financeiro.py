# -*- coding: utf-8 -*-
"""
Analisa a planilha de controle financeiro e retorna todas as linhas em que a
célula da coluna "OB" está com preenchimento AMARELO **e** o valor da célula
está vazio (sem número de OB preenchido).

Para cada ocorrência, extrai o código contido no final do campo "Descrição"
(ex: "ISS Retido MPI" -> código "MPI") e cruza com a TABELA_LOCALIDADE
(arquivo tabela_localidade.py): primeiro tenta encontrar correspondência
pela coluna "sigla"; se não encontrar, tenta pela coluna "org".

Retorna, para cada ocorrência:
    - Data
    - Descrição
    - Total OB
    - Código (extraído da Descrição)
    - CAT   (da tabela_localidade)
    - ORG   (da tabela_localidade)

Uso:
    python analisar_ob_amarelo.py [caminho_do_arquivo.xlsx] [--sheet NOME_DA_ABA]

Se nenhum caminho for informado, usa o arquivo padrão definido em DEFAULT_PATH.

Requisito: o arquivo tabela_localidade.py deve estar na mesma pasta deste
script (ou no PYTHONPATH).
"""

import re
import argparse
from datetime import datetime, date

import openpyxl

from tabela_localidade import TABELA_LOCALIDADE, buscar_por_sigla, buscar_por_org

DEFAULT_PATH = "controle-financeiro-2026-2.xlsx"

# Conjuntos de códigos conhecidos (em maiúsculo), extraídos da tabela de referência.
# Usados para reconhecer códigos compostos por mais de uma palavra (ex: "GOI 2").
_SIGLAS_CONHECIDAS = {
    item["sigla"].upper() for item in TABELA_LOCALIDADE if item.get("sigla")
}
_ORGS_CONHECIDOS = {
    item["org"].upper() for item in TABELA_LOCALIDADE if item.get("org")
}
_CODIGOS_CONHECIDOS = _SIGLAS_CONHECIDAS | _ORGS_CONHECIDOS

# Cor de referência para "amarelo". O Excel/LibreOffice normalmente grava
# amarelo puro como FFFF00 (com prefixo de alpha "FF" -> "FFFFFF00").
AMARELO_RGB = "FFFF00"


def _rgb_da_celula(cell):
    """Extrai o RGB (6 dígitos hex, sem alpha) do preenchimento de uma célula."""
    fill = cell.fill
    if fill is None or fill.patternType is None:
        return None
    fg = fill.fgColor
    if fg is None or fg.type != "rgb" or not fg.rgb:
        return None
    rgb = fg.rgb
    # rgb geralmente vem como "AARRGGBB" (8 dígitos); pegamos só "RRGGBB"
    return rgb[-6:].upper() if rgb else None


def _e_amarelo(cell, tolerancia=30):
    """
    Verifica se a célula está pintada de amarelo.
    Aceita pequenas variações de tom (tolerância nos componentes RGB),
    já que planilhas reais às vezes usam amarelos ligeiramente diferentes.
    """
    rgb = _rgb_da_celula(cell)
    if rgb is None:
        return False
    try:
        r, g, b = int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)
    except ValueError:
        return False

    alvo_r, alvo_g, alvo_b = 255, 255, 0
    return (
        abs(r - alvo_r) <= tolerancia
        and abs(g - alvo_g) <= tolerancia
        and b <= tolerancia
    )


def _formatar_data(valor):
    if isinstance(valor, (datetime, date)):
        return valor.strftime("%d/%m/%Y")
    return valor


def extrair_sigla(descricao):
    """
    Isola o código de localidade contido no final do texto de "Descrição".
    O código pode corresponder tanto à coluna "sigla" quanto à coluna "org"
    da tabela_localidade — a checagem de qual coluna bate é feita depois,
    em buscar_localidade().

    Estratégia:
    1. Tenta casar as duas últimas palavras contra os códigos conhecidos
       (sigla ou org combinados) — cobre casos compostos, ex: "GOI 2".
    2. Tenta casar apenas a última palavra contra os códigos conhecidos.
    3. Como último recurso, se a última palavra for só letras maiúsculas
       (2 a 5 caracteres), assume que é o código — mesmo que não exista
       na tabela (o lookup posterior indicará "não encontrado").
    """
    if not descricao:
        return None

    palavras = str(descricao).strip().split()
    if not palavras:
        return None

    # 1) duas últimas palavras (códigos compostos, ex: "GOI 2")
    if len(palavras) >= 2:
        candidata = " ".join(palavras[-2:]).upper()
        if candidata in _CODIGOS_CONHECIDOS:
            return candidata

    # 2) última palavra
    candidata = palavras[-1].upper()
    if candidata in _CODIGOS_CONHECIDOS:
        return candidata

    # 3) fallback: parece um código (só letras maiúsculas, 2 a 5 chars)?
    if re.fullmatch(r"[A-ZÀ-Ú]{2,5}", palavras[-1].upper()):
        return candidata

    return None


def buscar_localidade(codigo):
    """
    Busca o registro da tabela_localidade a partir de um código extraído da
    Descrição. Primeiro tenta casar contra a coluna "sigla"; se não achar,
    tenta contra a coluna "org".
    """
    if not codigo:
        return None
    registro = buscar_por_sigla(codigo)
    if registro is not None:
        return registro
    return buscar_por_org(codigo)


def encontrar_header(ws, colunas_esperadas=("OB", "Data", "Descrição", "Total OB")):
    """Localiza a linha de cabeçalho e o índice (1-based) de cada coluna esperada."""
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20)):
        valores = {str(c.value).strip(): c.column for c in row if c.value is not None}
        if all(col in valores for col in colunas_esperadas):
            linha_header = row[0].row
            return linha_header, valores
    raise ValueError(
        f"Não foi possível localizar uma linha de cabeçalho contendo as colunas: {colunas_esperadas}"
    )


def analisar(caminho_arquivo, nome_aba=None):
    wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
    ws = wb[nome_aba] if nome_aba else wb.active

    linha_header, col_idx = encontrar_header(ws)
    col_ob = col_idx["OB"]
    col_data = col_idx["Data"]
    col_desc = col_idx["Descrição"]
    col_total = col_idx["Total OB"]

    resultados = []
    for row in ws.iter_rows(min_row=linha_header + 1, max_row=ws.max_row):
        cell_ob = row[col_ob - 1]
        vazio = cell_ob.value is None or str(cell_ob.value).strip() == ""
        if vazio and _e_amarelo(cell_ob):
            descricao = row[col_desc - 1].value
            sigla = extrair_sigla(descricao)
            registro_localidade = buscar_localidade(sigla)

            resultados.append(
                {
                    "Data": _formatar_data(row[col_data - 1].value),
                    "Descrição": descricao,
                    "Total OB": row[col_total - 1].value,
                    "Código": sigla,
                    "CAT": registro_localidade["cat"] if registro_localidade else None,
                    "ORG": registro_localidade["org"] if registro_localidade else None,
                }
            )

    return resultados


def main():
    parser = argparse.ArgumentParser(description="Extrai ocorrências com OB pintada de amarelo.")
    parser.add_argument("arquivo", nargs="?", default=DEFAULT_PATH, help="Caminho do arquivo .xlsx")
    parser.add_argument("--sheet", default=None, help="Nome da aba (padrão: aba ativa)")
    args = parser.parse_args()

    resultados = analisar(args.arquivo, args.sheet)

    if not resultados:
        print("Nenhuma ocorrência com célula OB amarela foi encontrada.")
        return

    print(f"Total de ocorrências encontradas: {len(resultados)}\n")
    for i, item in enumerate(resultados, start=1):
        cat = item["CAT"] or "não encontrado"
        org = item["ORG"] or "não encontrado"
        print(
            f"{i}. Data: {item['Data']} | Descrição: {item['Descrição']} | "
            f"Total OB: {item['Total OB']} | Código: {item['Código']} | CAT: {cat} | ORG: {org}"
        )


if __name__ == "__main__":
    main()