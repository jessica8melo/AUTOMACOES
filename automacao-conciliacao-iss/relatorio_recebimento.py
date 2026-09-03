# -*- coding: utf-8 -*-
"""
Etapa 3: cruza o Relatório de Recebimento Integrado com os pagamentos
identificados nas Etapas 1 e 2.

Fluxo:
    1. Recebe a ORG (obtida na Etapa 1) e a lista de pagamentos daquela ORG
       (resultados da Etapa 2, cada item com pelo menos as chaves "NFF" e
       "Valor").
    2. Filtra, na planilha de Recebimento Integrado, todas as linhas cuja
       coluna "Organização" seja igual à ORG.
    3. Para cada linha filtrada, confere:
           N.F (Recebimento)  ==  NFF (Pagamento)     [comparação exata]
           ISS (Recebimento)  ==  Valor (Pagamento)   [comparação exata]
       Se ambos baterem: pinta a célula da coluna "N.F" de AMARELO e marca
       a linha como "OK".
       Se não baterem: NÃO pinta a célula e marca a linha como
       "INCONSISTENTE", com o motivo.

Observação sobre a planilha:
    A coluna "Alíquota" (L) contém fórmulas (=ISS/Valor Total da Nota*100).
    Este script não altera nenhuma dessas fórmulas nem os valores de que
    elas dependem — apenas o preenchimento (cor) da coluna "N.F". Ainda
    assim, como o arquivo é reescrito pelo openpyxl, o valor em cache das
    fórmulas é perdido até a planilha ser reaberta/recalculada no Excel ou
    LibreOffice (isso é uma limitação do próprio openpyxl, não um efeito
    deste script).

Uso (chamado a partir do main.py, um cruzamento por ORG):
    from relatorio_recebimento import verificar_por_org
    resultados = verificar_por_org(caminho, org, pagamentos_da_org)
"""

import argparse
import json

import openpyxl
from openpyxl.styles import PatternFill

DEFAULT_PATH = "Relatório Recebimento Integrado - 07-2026- ERP- Demais Filiais- Contas a Pagar.xlsx"

COL_ORGANIZACAO = "Organização"
COL_NF = "N.F"
COL_ISS = "ISS"
COL_FORNECEDOR = "Fornecedor"
COL_OBSERVACOES = "Observações"

AMARELO = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")


def _mapear_colunas(worksheet):
    """Lê a linha de cabeçalho (linha 1) e devolve {nome_da_coluna: índice}."""
    colunas = {}
    for cell in worksheet[1]:
        if cell.value:
            colunas[str(cell.value).strip()] = cell.column

    obrigatorias = [COL_ORGANIZACAO, COL_NF, COL_ISS, COL_FORNECEDOR]
    faltando = [c for c in obrigatorias if c not in colunas]
    if faltando:
        raise ValueError(
            "Coluna(s) não encontrada(s) na planilha de recebimento: "
            + ", ".join(faltando)
        )
    return colunas


def _observacao_da_linha(worksheet, linha, col_observacoes):
    """Devolve o texto da coluna 'Observações' dessa linha, ou None se vazio/coluna ausente."""
    if not col_observacoes:
        return None
    valor = worksheet.cell(row=linha, column=col_observacoes).value
    if valor is None:
        return None
    texto = str(valor).strip()
    return texto or None


def _comentarios_da_linha(worksheet, linha, colunas):
    """Devolve os comentários nativos do Excel (ícone de nota) anexados a células dessa linha."""
    comentarios = []
    for nome_coluna, col in colunas.items():
        celula = worksheet.cell(row=linha, column=col)
        if celula.comment is not None and celula.comment.text:
            comentarios.append(f"{nome_coluna}: {celula.comment.text.strip()}")
    return comentarios


def _avisos_da_linha(worksheet, linha, colunas, col_observacoes):
    """
    Reúne, em uma lista de textos, os avisos encontrados nessa linha:
    texto na coluna 'Observações' e/ou comentários nativos do Excel 
    em qualquer célula da linha.
    """
    avisos = []
    observacao = _observacao_da_linha(worksheet, linha, col_observacoes)
    if observacao:
        avisos.append(f'Observações: "{observacao}"')

    for comentario in _comentarios_da_linha(worksheet, linha, colunas):
        avisos.append(f"comentário em {comentario}")

    return avisos


def _para_texto(valor):
    """
    Converte um valor para texto de forma EXATA (sem normalizar formatação
    numérica): 780 vira '780', 780.0 vira '780.0', 'ABC ' vira 'ABC ' (com
    strip apenas nas pontas).

    Ajuste esta função se, na prática, "N.F" e "NFF" vierem em formatos
    diferentes (ex.: '780' vs '780.00') e for necessário normalizar antes
    de comparar.
    """
    if valor is None:
        return ""
    return str(valor).strip()


def filtrar_por_org(caminho, org, sheet=None):
    """
    Retorna, sem modificar nada, as linhas da planilha de recebimento cuja
    coluna "Organização" seja igual a `org`. Cada item tem:
        {"Fornecedor", "N.F", "ISS", "linha", "avisos"}
    "avisos" é uma lista com textos indicando, por exemplo, que a linha tem 
    texto na coluna "Observações"/comentário.
    Útil para inspecionar os dados antes de rodar verificar_por_org().
    """
    workbook = openpyxl.load_workbook(caminho, data_only=True)
    worksheet = workbook[sheet] if sheet else workbook.active
    colunas = _mapear_colunas(worksheet)

    col_org = colunas[COL_ORGANIZACAO]
    col_nf = colunas[COL_NF]
    col_iss = colunas[COL_ISS]
    col_fornecedor = colunas[COL_FORNECEDOR]
    col_observacoes = colunas.get(COL_OBSERVACOES)

    encontrados = []
    for linha in range(2, worksheet.max_row + 1):
        valor_org = worksheet.cell(row=linha, column=col_org).value
        if valor_org is None or str(valor_org).strip() != str(org).strip():
            continue
        encontrados.append({
            "Fornecedor": worksheet.cell(row=linha, column=col_fornecedor).value,
            "N.F": worksheet.cell(row=linha, column=col_nf).value,
            "ISS": worksheet.cell(row=linha, column=col_iss).value,
            "linha": linha,
            "avisos": _avisos_da_linha(worksheet, linha, colunas, col_observacoes),
        })
    return encontrados


def verificar_por_org(caminho, org, pagamentos, sheet=None, salvar=True):
    """
    Filtra a planilha de Recebimento Integrado pela coluna "Organização" ==
    `org` e confere cada linha encontrada contra `pagamentos` (resultados
    da Etapa 2 para essa ORG; cada item precisa ter pelo menos as chaves
    "NFF" e "Valor").

    Critério de "tudo certo" (comparação exata em ambos):
        N.F (recebimento)  ==  NFF (pagamento)
        ISS (recebimento)  ==  Valor (pagamento)

    Se baterem: pinta a célula "N.F" da linha de AMARELO e marca "OK".
    Se não baterem: não pinta e marca "INCONSISTENTE" com o motivo.

    Cada NFF de `pagamentos` só é usada uma única vez (para não casar duas
    linhas do recebimento com o mesmo pagamento).

    Além disso, para cada linha processada, verifica se tem texto na coluna "Observações" 
    (ou comentário nativo do Excel em alguma célula da linha). Se houver, imprime um aviso 
    no terminal na hora e inclui os detalhes no campo "avisos" do resultado.

    Retorna uma lista de dicts:
        {
            "Organização", "Fornecedor", "N.F", "ISS", "linha",
            "status": "OK" | "INCONSISTENTE",
            "motivo": None | texto explicando a inconsistência,
            "avisos": lista de textos (vazia se não houver nada a avisar),
        }
    """
    workbook = openpyxl.load_workbook(caminho)
    worksheet = workbook[sheet] if sheet else workbook.active
    colunas = _mapear_colunas(worksheet)

    col_org = colunas[COL_ORGANIZACAO]
    col_nf = colunas[COL_NF]
    col_iss = colunas[COL_ISS]
    col_fornecedor = colunas[COL_FORNECEDOR]
    col_observacoes = colunas.get(COL_OBSERVACOES)

    # Índice de pagamentos por NFF (texto exato); cada NFF pode aparecer
    # mais de uma vez, então guardamos uma lista e vamos "consumindo" os
    # itens conforme casam com linhas do recebimento.
    pagamentos_por_nff = {}
    for item in pagamentos:
        chave = _para_texto(item.get("NFF"))
        pagamentos_por_nff.setdefault(chave, []).append(item)

    resultados = []
    for linha in range(2, worksheet.max_row + 1):
        valor_org = worksheet.cell(row=linha, column=col_org).value
        if valor_org is None or str(valor_org).strip() != str(org).strip():
            continue

        cel_nf = worksheet.cell(row=linha, column=col_nf)
        cel_iss = worksheet.cell(row=linha, column=col_iss)
        cel_fornecedor = worksheet.cell(row=linha, column=col_fornecedor)

        nf_texto = _para_texto(cel_nf.value)
        iss_texto = _para_texto(cel_iss.value)
        candidatos = pagamentos_por_nff.get(nf_texto, [])

        pagamento_usado = None
        for candidato in candidatos:
            if iss_texto == _para_texto(candidato.get("Valor")):
                pagamento_usado = candidato
                break

        if pagamento_usado is not None:
            status = "OK"
            motivo = None
            cel_nf.fill = AMARELO
            candidatos.remove(pagamento_usado)  # evita reusar o mesmo pagamento
        elif candidatos:
            status = "INCONSISTENTE"
            motivo = (
                f"NFF {nf_texto} encontrada na Etapa 2, mas ISS ({iss_texto}) "
                f"não bate com o Valor de nenhum pagamento associado a essa NFF."
            )
        else:
            status = "INCONSISTENTE"
            motivo = f"Nenhum pagamento com NFF {nf_texto} foi encontrado na Etapa 2."

        avisos = _avisos_da_linha(worksheet, linha, colunas, col_observacoes)
        if avisos:
            print(
                f"⚠️  Atenção — linha {linha} (Fornecedor: {cel_fornecedor.value}, "
                f"N.F: {cel_nf.value}): {'; '.join(avisos)}"
            )

        resultados.append({
            "Organização": valor_org,
            "Fornecedor": cel_fornecedor.value,
            "N.F": cel_nf.value,
            "ISS": cel_iss.value,
            "linha": linha,
            "status": status,
            "motivo": motivo,
            "avisos": avisos,
        })

    if salvar:
        workbook.save(caminho)

    return resultados


def imprimir_resultados(org, resultados):
    print(f"\n--- ORG: {org} ---")
    if not resultados:
        print("Nenhuma linha de Recebimento encontrada para essa ORG.")
        return

    ok = sum(1 for r in resultados if r["status"] == "OK")
    inconsistentes = len(resultados) - ok
    com_aviso = sum(1 for r in resultados if r.get("avisos"))
    resumo = f"Total de linhas: {len(resultados)}  |  OK: {ok}  |  Inconsistentes: {inconsistentes}"
    if com_aviso:
        resumo += f"  |  Com aviso: {com_aviso}"
    print(resumo + "\n")

    for i, item in enumerate(resultados, start=1):
        linha_info = (
            f"{i}. [{item['status']}] Linha {item['linha']} | "
            f"Fornecedor: {item['Fornecedor']} | N.F: {item['N.F']} | ISS: {item['ISS']}"
        )
        print(linha_info)
        if item["motivo"]:
            print(f"   Motivo: {item['motivo']}")
        if item.get("avisos"):
            print(f"   ⚠️  Aviso: {'; '.join(item['avisos'])}")


def _main():
    parser = argparse.ArgumentParser(
        description=(
            "Etapa 3 — cruza o Relatório de Recebimento Integrado (coluna "
            "Organização) com os pagamentos de uma ORG (JSON com NFF e "
            "Valor), pintando de amarelo a célula N.F quando tudo bate."
        )
    )
    parser.add_argument("--recebimento", default=DEFAULT_PATH,
                         help="Caminho do arquivo .xlsx de Recebimento Integrado")
    parser.add_argument("--sheet-recebimento", default=None,
                         help="Nome da aba do recebimento (padrão: aba ativa)")
    parser.add_argument("--org", required=True, help="ORG a filtrar (ex.: LAU)")
    parser.add_argument(
        "--pagamentos-json", required=True,
        help=(
            'Caminho de um .json com a lista de pagamentos dessa ORG, '
            'ex.: [{"NFF": "996.02", "Valor": 21603.16}, ...] '
            '(mesmo formato dos resultados da Etapa 2)'
        )
    )
    parser.add_argument("--nao-salvar", action="store_true",
                         help="Não grava as cores no arquivo (só mostra o resultado)")
    args = parser.parse_args()

    with open(args.pagamentos_json, encoding="utf-8") as f:
        pagamentos = json.load(f)

    resultados = verificar_por_org(
        args.recebimento, args.org, pagamentos,
        sheet=args.sheet_recebimento, salvar=not args.nao_salvar,
    )
    imprimir_resultados(args.org, resultados)


if __name__ == "__main__":
    _main()