"""
Extrator de extrato de conta corrente do Banco do Brasil (PDF -> dados estruturados)
=====================================================================================

Lê um PDF (ou uma pasta com vários PDFs) de "Consultas - Extrato de conta corrente"
do BB e extrai:
  - Agência, número da conta e titular
  - Dia (ou período) do extrato
  - Lançamentos: histórico, documento, valor e se é Recebimento (C) ou Pagamento (D)

A extração usa as TABELAS do PDF (pdfplumber), não o texto corrido — isso deixa o
parsing muito mais confiável, inclusive em extratos com muitas linhas e múltiplas
páginas.

Uso:
    # um único PDF
    python extrato_bb_parser.py extrato.pdf

    # uma pasta inteira (varre todos os .pdf, inclusive em subpastas)
    python extrato_bb_parser.py --dir caminho/da/pasta

    # salvando os resultados
    python extrato_bb_parser.py --dir caminho/da/pasta --csv saida.csv --json saida.json

    # gerando planilha organizada por conta (Pagamentos / Recebimentos separados)
    python extrato_bb_parser.py --dir caminho/da/pasta --xlsx saida.xlsx
"""

import argparse
import csv
import json
import re
import sys
from datetime import datetime
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

DATE_RE = re.compile(r"^\d{2}/\d{2}/\d{4}$")
LOTE_PREFIX_RE = re.compile(r"^\d{2,4}\s+")

# Larguras de colunas da tabela "Lançamentos" no layout do BB:
# Dt.balancete | Dt.movimento | Ag.origem | Lote | Histórico | Documento | Valor R$ | Saldo
N_COLS_LANCAMENTOS = 8


def _limpar(valor):
    if valor is None:
        return ""
    return re.sub(r"\s+", " ", valor.replace("\n", " ")).strip()


def _extrair_linhas_tabela(pdf_path: str):
    """Percorre todas as páginas do PDF e junta as linhas da tabela de lançamentos
    (ela pode continuar em várias páginas, sem repetir o cabeçalho)."""
    linhas = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for tabela in page.extract_tables():
                if not tabela or len(tabela[0]) != N_COLS_LANCAMENTOS:
                    continue
                for row in tabela:
                    primeira = _limpar(row[0])
                    if primeira.startswith("Dt."):
                        continue  # linha de cabeçalho
                    linhas.append(row)
    return linhas


def _cabecalho(pdf_path: str) -> dict:
    with pdfplumber.open(pdf_path) as pdf:
        texto = "\n".join(p.extract_text() or "" for p in pdf.pages)
    linhas_txt = [l.strip() for l in texto.split("\n")]

    agencia = conta_raw = periodo = None
    for l in linhas_txt:
        if l.startswith("Agência"):
            agencia = l.replace("Agência", "").strip()
        elif l.startswith("Conta corrente"):
            conta_raw = l.replace("Conta corrente", "").strip()
        elif l.startswith("Período do extrato"):
            periodo = l.replace("Período do extrato", "").strip()

    # Formato comum: "200000-8BB T SERVICOS S.A." (sem espaço entre o dígito
    # verificador da conta e o nome do titular) ou, mais raramente, com espaço.
    m = re.match(r"^([\d.]+-[0-9A-Za-z])\s*(.*)$", conta_raw or "")
    numero_conta = m.group(1) if m else conta_raw
    titular = m.group(2).strip() if m and m.group(2) else None

    datas = re.findall(r"\d{2}\s*/\s*\d{2}\s*/\s*\d{4}", periodo or "")
    datas = [d.replace(" ", "") for d in datas]
    dia_extrato = datas[0] if len(datas) == 2 and datas[0] == datas[1] else None

    return {
        "agencia": agencia,
        "numero_conta": numero_conta,
        "titular": titular,
        "dia_extrato": dia_extrato,
        "periodo": periodo,
    }


def _deve_ignorar(historico: str) -> bool:
    """Identifica linhas que não são lançamentos de fato (saldo, saldo anterior, agência)."""
    h = (historico or "").strip().upper()
    return (
        h.startswith("SALDO")
        or h.startswith("AGÊNCIA")
        or h.startswith("AGENCIA")
    )


def parse_extrato(pdf_path: str) -> dict:
    """Lê um PDF de extrato BB e devolve um dicionário estruturado."""
    dados = _cabecalho(pdf_path)
    linhas = _extrair_linhas_tabela(pdf_path)

    lancamentos = []
    for row in linhas:
        dt_balancete = _limpar(row[0])

        if dt_balancete and DATE_RE.match(dt_balancete):
            historico = LOTE_PREFIX_RE.sub("", _limpar(row[4]))
            documento = _limpar(row[5]) or None

            valor_col = _limpar(row[6])   # "Valor R$" -> lançamentos normais
            saldo_col = _limpar(row[7])   # "Saldo"    -> saldo anterior / saldo do dia
            bruto = valor_col or saldo_col

            tokens = bruto.split()
            tipo = tokens[-1] if tokens and tokens[-1] in ("C", "D") else None
            valor = tokens[-2] if tipo and len(tokens) >= 2 else (tokens[0] if tokens else None)

            tipo_registro = "lancamento" if documento else "saldo"

            if tipo_registro == "saldo" or _deve_ignorar(historico):
                continue  # ignora linhas de SALDO / Saldo anterior / Agência

            lancamentos.append({
                "data": dt_balancete,
                "historico": historico,
                "documento": documento,
                "valor": valor,
                "natureza": ("Recebimento" if tipo == "C" else "Pagamento") if tipo else None,
                "tipo_registro": tipo_registro,
            })
        else:
            # linha de continuação: acrescenta ao histórico do lançamento anterior
            complemento = _limpar(row[4])
            if complemento and lancamentos:
                lancamentos[-1]["historico"] = (lancamentos[-1]["historico"] + " " + complemento).strip()

    dados["lancamentos"] = lancamentos
    dados["arquivo"] = Path(pdf_path).name
    return dados


def parse_pasta(dir_path: str, recursivo: bool = True) -> list:
    """Lê todos os PDFs de uma pasta e devolve uma lista de dicionários (um por arquivo)."""
    pasta = Path(dir_path)
    padrao = "**/*.pdf" if recursivo else "*.pdf"
    arquivos = sorted(
        p for p in pasta.glob(padrao)
        if not p.name.startswith("._") and "__MACOSX" not in p.parts
    )

    if not arquivos:
        print(f"Nenhum PDF encontrado em: {dir_path}", file=sys.stderr)

    resultados = []
    for arq in arquivos:
        try:
            resultados.append(parse_extrato(str(arq)))
        except Exception as e:
            print(f"Falha ao processar '{arq.name}': {e}", file=sys.stderr)
    return resultados


def salvar_csv(lista_extratos: list, caminho: str) -> None:
    campos = ["arquivo", "agencia", "numero_conta", "dia_extrato",
              "data", "historico", "documento", "valor", "natureza", "tipo_registro"]
    with open(caminho, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=campos)
        writer.writeheader()
        for extrato in lista_extratos:
            for lc in extrato["lancamentos"]:
                writer.writerow({
                    "arquivo": extrato["arquivo"],
                    "agencia": extrato["agencia"],
                    "numero_conta": extrato["numero_conta"],
                    "dia_extrato": extrato["dia_extrato"],
                    **lc,
                })


def imprimir_resumo(dados: dict) -> None:
    print(f"\nArquivo: {dados['arquivo']}")
    print(f"Agência: {dados['agencia']}  |  Conta: {dados['numero_conta']}"
          + (f" ({dados['titular']})" if dados.get("titular") else ""))
    print(f"Dia do extrato: {dados['dia_extrato'] or dados['periodo']}")
    print("-" * 90)
    for lc in dados["lancamentos"]:
        marca = "[SALDO]" if lc["tipo_registro"] == "saldo" else f"[{lc['natureza']}]"
        doc = lc["documento"] or "-"
        print(f"{lc['data']}  {marca:<14} R$ {lc['valor']:<12} doc: {doc:<22} {lc['historico']}")
    print("-" * 90)


FONTE = "Arial"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONTE, bold=True, color="FFFFFF", size=11)
TITULO_FONT = Font(name=FONTE, bold=True, size=14, color="1F4E78")
SUBTITULO_FONT = Font(name=FONTE, size=10, color="404040")
SECAO_PAG_FILL = PatternFill("solid", fgColor="C00000")
SECAO_REC_FILL = PatternFill("solid", fgColor="1E7B34")
SECAO_FONT = Font(name=FONTE, bold=True, size=12, color="FFFFFF")
TOTAL_FONT = Font(name=FONTE, bold=True, size=10)
TOTAL_FILL = PatternFill("solid", fgColor="D9D9D9")
BORDA_FINA = Border(bottom=Side(style="thin", color="BFBFBF"))
COLUNAS_XLSX = ["Data", "Histórico", "Documento", "Valor (R$)"]


def _valor_para_float(valor_str):
    if not valor_str:
        return 0.0
    limpo = valor_str.replace(".", "").replace(",", ".")
    try:
        return float(limpo)
    except ValueError:
        return 0.0


def _data_para_ordenacao(data_str):
    try:
        return datetime.strptime(data_str, "%d/%m/%Y")
    except (ValueError, TypeError):
        return datetime.min


def _nome_aba(numero_conta: str, usados: set) -> str:
    nome = re.sub(r'[:\\/?*\[\]]', "-", numero_conta or "Conta")[:31]
    base = nome
    i = 2
    while nome in usados:
        sufixo = f" ({i})"
        nome = base[: 31 - len(sufixo)] + sufixo
        i += 1
    usados.add(nome)
    return nome


def _escrever_secao(ws, linha, titulo, fill, lancamentos):
    """Escreve o cabeçalho da seção + tabela de lançamentos. Retorna a próxima linha livre."""
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=4)
    cel = ws.cell(row=linha, column=1, value=titulo)
    cel.font = SECAO_FONT
    cel.fill = fill
    cel.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[linha].height = 22
    linha += 1

    for col, nome_col in enumerate(COLUNAS_XLSX, start=1):
        c = ws.cell(row=linha, column=col, value=nome_col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center" if col != 2 else "left", vertical="center")
    linha += 1

    total = 0.0
    for lc in sorted(lancamentos, key=lambda x: _data_para_ordenacao(x["data"])):
        ws.cell(row=linha, column=1, value=lc["data"]).alignment = Alignment(horizontal="center")
        ws.cell(row=linha, column=2, value=lc["historico"])
        ws.cell(row=linha, column=3, value=lc["documento"] or "-").alignment = Alignment(horizontal="center")
        v = _valor_para_float(lc["valor"])
        total += v
        vc = ws.cell(row=linha, column=4, value=v)
        vc.number_format = '#,##0.00'
        vc.alignment = Alignment(horizontal="right")
        for col in range(1, 5):
            ws.cell(row=linha, column=col).font = Font(name=FONTE, size=10)
            ws.cell(row=linha, column=col).border = BORDA_FINA
        linha += 1

    if not lancamentos:
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=4)
        c = ws.cell(row=linha, column=1, value="(nenhum lançamento)")
        c.font = Font(name=FONTE, italic=True, size=10, color="808080")
        c.alignment = Alignment(horizontal="center")
        linha += 1

    # linha de total
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=3)
    tc = ws.cell(row=linha, column=1, value=f"Total ({len(lancamentos)} lançamento(s))")
    tc.font = TOTAL_FONT
    tc.fill = TOTAL_FILL
    tc.alignment = Alignment(horizontal="right", indent=1)
    tvc = ws.cell(row=linha, column=4, value=total)
    tvc.number_format = '#,##0.00'
    tvc.font = TOTAL_FONT
    tvc.fill = TOTAL_FILL
    tvc.alignment = Alignment(horizontal="right")
    for col in range(1, 5):
        ws.cell(row=linha, column=col).fill = TOTAL_FILL
    linha += 1

    return linha + 1  # deixa uma linha em branco depois da seção


def salvar_xlsx(lista_extratos: list, caminho: str) -> None:
    """Gera uma planilha com uma aba por conta, cada uma dividida em
    seção de Pagamentos e seção de Recebimentos."""
    if not lista_extratos:
        raise ValueError("Nenhum extrato para gerar a planilha.")

    # Agrupa lançamentos por conta (chave = numero_conta)
    contas = {}
    for extrato in lista_extratos:
        chave = extrato["numero_conta"] or "Sem número"
        conta = contas.setdefault(chave, {
            "agencia": extrato["agencia"],
            "titular": extrato["titular"],
            "lancamentos": [],
            "dias": set(),
        })
        conta["lancamentos"].extend(extrato["lancamentos"])
        if extrato.get("dia_extrato"):
            conta["dias"].add(extrato["dia_extrato"])

    wb = Workbook()
    wb.remove(wb.active)
    abas_usadas = set()

    for numero_conta in sorted(contas.keys()):
        conta = contas[numero_conta]
        ws = wb.create_sheet(_nome_aba(numero_conta, abas_usadas))

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 55
        ws.column_dimensions["C"].width = 24
        ws.column_dimensions["D"].width = 16

        linha = 1
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=4)
        c = ws.cell(row=linha, column=1, value=f"Conta {numero_conta}")
        c.font = TITULO_FONT
        linha += 1

        subtitulo = f"Agência {conta['agencia'] or '-'}"
        if conta["titular"]:
            subtitulo += f"  |  {conta['titular']}"
        if conta["dias"]:
            periodo_fmt = ", ".join(sorted(conta["dias"], key=_data_para_ordenacao))
            subtitulo += f"  |  Dias: {periodo_fmt}"
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=4)
        c = ws.cell(row=linha, column=1, value=subtitulo)
        c.font = SUBTITULO_FONT
        linha += 2

        pagamentos = [lc for lc in conta["lancamentos"] if lc["natureza"] == "Pagamento"]
        recebimentos = [lc for lc in conta["lancamentos"] if lc["natureza"] == "Recebimento"]

        linha = _escrever_secao(ws, linha, f"PAGAMENTOS ({len(pagamentos)})", SECAO_PAG_FILL, pagamentos)
        linha = _escrever_secao(ws, linha, f"RECEBIMENTOS ({len(recebimentos)})", SECAO_REC_FILL, recebimentos)

        ws.freeze_panes = "A5"

    wb.save(caminho)


def main():
    parser = argparse.ArgumentParser(description="Extrai lançamentos de extrato(s) BB em PDF")
    grupo = parser.add_mutually_exclusive_group(required=True)
    grupo.add_argument("pdf", nargs="?", help="Caminho de um único arquivo PDF")
    grupo.add_argument("--dir", help="Caminho de uma pasta com vários PDFs")
    parser.add_argument("--sem-recursivo", action="store_true",
                         help="Com --dir, não entra em subpastas (por padrão entra)")
    parser.add_argument("--csv", help="Caminho para salvar os lançamentos em CSV", default=None)
    parser.add_argument("--json", help="Caminho para salvar os dados completos em JSON", default=None)
    parser.add_argument("--xlsx", help="Caminho para salvar planilha organizada por conta "
                                        "(Pagamentos e Recebimentos separados)", default=None)
    parser.add_argument("--silencioso", action="store_true", help="Não imprime o resumo no terminal")
    args = parser.parse_args()

    if args.dir:
        if not Path(args.dir).is_dir():
            print(f"Pasta não encontrada: {args.dir}", file=sys.stderr)
            sys.exit(1)
        resultados = parse_pasta(args.dir, recursivo=not args.sem_recursivo)
    else:
        if not Path(args.pdf).exists():
            print(f"Arquivo não encontrado: {args.pdf}", file=sys.stderr)
            sys.exit(1)
        resultados = [parse_extrato(args.pdf)]

    if not args.silencioso:
        for dados in resultados:
            imprimir_resumo(dados)
        total = sum(len(d["lancamentos"]) for d in resultados)
        print(f"\n{len(resultados)} arquivo(s) processado(s), {total} linha(s) extraída(s) no total.")

    if args.csv:
        salvar_csv(resultados, args.csv)
        print(f"\nCSV salvo em: {args.csv}")

    if args.json:
        with open(args.json, "w", encoding="utf-8") as f:
            json.dump(resultados, f, ensure_ascii=False, indent=2)
        print(f"JSON salvo em: {args.json}")

    if args.xlsx:
        salvar_xlsx(resultados, args.xlsx)
        print(f"Planilha (xlsx) salva em: {args.xlsx}")


if __name__ == "__main__":
    main()