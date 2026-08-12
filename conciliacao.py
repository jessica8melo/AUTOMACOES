"""
Passo 2: em cada aba, concilia pagamento(s) com OB dentro dessa mesma aba.
A busca por match segue uma ordem de prioridade:

    1) Match exato de UM único pagamento cujo valor bate com o de UMA OB.
       Feito primeiro, pra TODAS as OBs, sem nenhuma ambiguidade (não
       depende de olhar combinações).

    2) Entre as OBs que sobraram sem match no passo 1, tenta somar 2 ou
       mais delas entre si (ex.: duas remessas que na prática foram pagas
       juntas em um único lote) e ver se essa soma bate com um pagamento ou
       combinação de pagamentos ainda disponível - ver
       `_buscar_combo_obs_pagamentos`. Importante isso vir ANTES do passo
       3: se uma OB isolada "coubesse" dentro de um pool de pagamentos de
       valor repetido (ex.: vários pagamentos de R$300,00 iguais), o passo
       3 casaria essa OB sozinha com um subconjunto arbitrário desse pool,
       "gastando" pagamentos que na verdade deveriam ficar reservados para
       serem somados junto com OUTRA OB (ex.: duas OBs que juntas fecham
       o valor de TODO o pool de repetidos). Rodar a combinação de OBs
       primeiro evita esse tipo de match "ganancioso" e incompleto.

    3) Para as OBs que ainda sobrarem, combinações de 2 ou mais pagamentos
       cuja soma bate com o valor de UMA OB isolada, dando preferência para
       pagamentos que estejam fisicamente próximos entre si na planilha
       (idealmente na linha logo abaixo ou logo acima um do outro).

    4) Se nada acima servir, vale qualquer combinação exata encontrada pela
       busca exaustiva (com poda), mesmo que os pagamentos estejam
       espalhados/longe uns dos outros.

Quando acha, pinta a(s) linha(s) do(s) pagamento(s) e a(s) linha(s) da(s)
OB(s) com a mesma cor, deixando visualmente claro quem está atrelado a
quem (inclusive quando o grupo tem mais de uma OB).

Abas sem bloco de Pagamentos e/ou sem tabela OB/VALOR são ignoradas (nada é
alterado nelas).
"""

import itertools
import math

import openpyxl
from openpyxl.styles import PatternFill

from estilos import CORES_CONCILIACAO, NORMAL_FONT, NUM_COLS_PRINCIPAL
from helpers_planilha import localizar_titulo, localizar_bloco_ob, _centavos, _normalizar_abas


def _linhas_pagamentos(ws, linha_titulo_pag):
    """Lê a tabela principal de Pagamentos (colunas A:G) dessa aba,
    começando logo após o título 'PAGAMENTOS (N)' e o cabeçalho de colunas
    que vem em seguida. Devolve [(linha, quantia)] enquanto a coluna
    'Linha' (A) tiver valor."""
    linhas = []
    r = linha_titulo_pag + 2  # +1 = cabeçalho de colunas, +2 = primeira linha de dado
    while ws.cell(row=r, column=1).value is not None:
        quantia = ws.cell(row=r, column=5).value  # coluna E = Quantia
        if quantia is not None:
            linhas.append((r, quantia))
        r += 1
    return linhas


def _linhas_ob_com_linha(ws, col_ob, col_valor, linha_inicial):
    """Como _ler_linhas_ob, mas guardando também o número da linha (usado
    pra pintar depois). Também para de ler se a coluna VALOR trouxer algo
    que não seja número (ver comentário em _ler_linhas_ob)."""
    linhas = []
    r = linha_inicial
    while True:
        v_ob = ws.cell(row=r, column=col_ob).value
        v_valor = ws.cell(row=r, column=col_valor).value
        if v_ob is None and v_valor is None:
            break
        if v_valor is not None and not isinstance(v_valor, (int, float)):
            break
        if v_valor is not None:
            linhas.append((r, v_ob, v_valor))
        r += 1
    return linhas


def _score_proximidade(combo):
    """Gera uma penalidade para combinações de 2+ pagamentos cujas linhas
    estão mais distantes entre si na planilha. Menor score é melhor. Dá
    peso extra para sequências contíguas (linhas vizinhas), que é o que
    queremos priorizar quando um único pagamento não bate exato com a OB.

    Usada apenas para desempate ENTRE combinações de 2+ itens - o caso de
    um único pagamento com valor exato é tratado à parte, antes disso, em
    `_buscar_subconjunto`."""
    linhas = sorted(linha for linha, _ in combo)
    span = linhas[-1] - linhas[0]
    gaps = [b - a for a, b in zip(linhas, linhas[1:])]
    contiguas = sum(1 for gap in gaps if gap == 1)
    return (span, sum(gaps), -contiguas, linhas[0])


def _particionar_combo_por_ob(combo_obs, combo_pagamentos, max_tamanho):
    """Quando um grupo junta 2+ OBs vs uma combinação de pagamentos, tenta
    achar, dentro desses mesmos pagamentos, qual subconjunto bate exato com
    o valor de CADA OB individualmente (ex.: OB A = pagamentos 1+2, OB B =
    pagamentos 3+4). Usado só pra decidir a cor de cada linha na hora de
    pintar - não muda a conciliação em si.

    Vai das maiores OBs pras menores, indo tentando tirar do pool comum um
    subconjunto exato pra cada uma (via `_buscar_subconjunto`); a última OB
    fica com o que sobrar, sem outra busca, e só validando que a soma bate.

    Devolve uma lista de dicts [{"ob_row", "ob_codigo", "valor",
    "pagamento_rows"}, ...], um por OB, ou None se não achar uma divisão
    exata (nesse caso quem chamou deve cair de volta pra pintar o grupo
    inteiro com uma cor só)."""
    obs_ordenadas = sorted(combo_obs, key=lambda x: x[2], reverse=True)
    pool = list(combo_pagamentos)
    particoes = []

    for ob_row, ob_codigo, valor in obs_ordenadas[:-1]:
        alvo = _centavos(valor)
        combo = _buscar_subconjunto(pool, alvo, max_tamanho)
        if combo is None:
            return None
        particoes.append({
            "ob_row": ob_row,
            "ob_codigo": ob_codigo,
            "valor": valor,
            "pagamento_rows": [linha for linha, _ in combo],
        })
        usados = set(linha for linha, _ in combo)
        pool = [p for p in pool if p[0] not in usados]

    # última OB fica com o restante - só confirma que a soma bate certinho
    ultima_ob_row, ultima_ob_codigo, ultima_valor = obs_ordenadas[-1]
    if not pool or _centavos(ultima_valor) != sum(_centavos(q) for _, q in pool):
        return None
    particoes.append({
        "ob_row": ultima_ob_row,
        "ob_codigo": ultima_ob_codigo,
        "valor": ultima_valor,
        "pagamento_rows": [linha for linha, _ in pool],
    })

    return particoes


def _buscar_subconjunto(pagamentos_disponiveis, alvo_centavos, max_tamanho):
    """Procura pagamento(s) cuja soma bate exatamente com `alvo_centavos`,
    seguindo esta ordem de prioridade:

        1) Match exato de UM único pagamento (retorna na hora, sem olhar
           para combinações de 2+ itens).
        2) Combinações de 2 ou mais pagamentos, priorizando as que usam
           pagamentos mais próximos entre si na planilha (idealmente linhas
           vizinhas / contíguas) - ver `_score_proximidade`.
        3) Se não houver combinação "próxima", qualquer combinação exata
           encontrada pela busca com poda serve (tentativa e erro geral,
           sem preferência de posição).

    Devolve a combinação (tupla de (linha, quantia)) ou None se nada bater.

    A busca em si usa poda (branch and bound) em vez de testar todas as
    combinações via itertools.combinations: os pagamentos são ordenados por
    valor, e a recursão corta um ramo assim que a soma parcial já passa do
    alvo (como a lista está em ordem crescente, os itens seguintes só
    aumentam a soma, então não há por que continuar). Sem isso, uma aba com
    ~50-80 pagamentos e várias OBs sem correspondência podia levar minutos
    (ou até travar na prática) tentando C(n, 6) combinações repetidamente."""
    itens = sorted(pagamentos_disponiveis, key=lambda x: _centavos(x[1]))
    valores = [_centavos(q) for _, q in itens]
    n = len(itens)
    max_tamanho = min(max_tamanho, n)

    # --- Prioridade 1: pagamento único com valor exatamente igual ao da OB ---
    for item, valor_centavos in zip(itens, valores):
        if valor_centavos == alvo_centavos:
            return (item,)

    # --- Prioridades 2 e 3: combinações de 2+ itens, do menor pro maior
    #     tamanho, com desempate por proximidade entre linhas ---
    melhor_combo = None
    melhor_score = None

    for tamanho in range(2, max_tamanho + 1):
        combo = _buscar_tamanho_fixo(itens, valores, alvo_centavos, tamanho)
        if combo is None:
            continue

        score = _score_proximidade(combo)
        if melhor_combo is None or score < melhor_score:
            melhor_combo = combo
            melhor_score = score

    return melhor_combo


def _buscar_tamanho_fixo(itens, valores, alvo, tamanho):
    """Busca com poda por uma combinação de exatamente `tamanho` itens (já
    ordenados por valor crescente em `itens`/`valores`) cuja soma bata com
    `alvo`. Devolve a tupla de itens ou None. Quando houver múltiplas
    combinações exatas desse mesmo tamanho, a melhor é escolhida pelo
    critério de proximidade entre linhas (`_score_proximidade`)."""
    n = len(itens)
    escolhidos = [0] * tamanho
    melhor_combo = None

    def rec(inicio, k, soma_parcial):
        nonlocal melhor_combo
        if k == tamanho:
            if soma_parcial == alvo:
                combo = tuple(itens[i] for i in escolhidos)
                if melhor_combo is None:
                    melhor_combo = combo
                else:
                    atual_score = _score_proximidade(combo)
                    melhor_score = _score_proximidade(melhor_combo)
                    if atual_score < melhor_score:
                        melhor_combo = combo
            return
        # não vale a pena tentar índices onde não sobram itens suficientes
        # para completar os `tamanho - k` restantes
        limite = n - (tamanho - k)
        for i in range(inicio, limite + 1):
            nova_soma = soma_parcial + valores[i]
            if nova_soma > alvo:
                # itens[i:] só tem valores >= valores[i] (lista ordenada),
                # então nenhum item a partir daqui pode mais servir
                break
            escolhidos[k] = i
            rec(i + 1, k + 1, nova_soma)
        return None

    rec(0, 0, 0)
    return melhor_combo


# trava de segurança: para um dado tamanho de combinação, se o número de
# combinações possíveis (n escolhe k) passar disso, esse tamanho é pulado em
# vez de arriscar travar o script numa aba com muitos itens sem match
LIMITE_COMBINACOES = 200_000


def _enumerar_somas(itens, valor_fn, max_tamanho, teto_centavos=None, limite=LIMITE_COMBINACOES):
    """Enumera, com poda, combinações de `itens` de tamanho 1 até
    `max_tamanho`, devolvendo um mapa {soma_centavos: combinação}. `valor_fn`
    extrai o valor monetário de um item. Os itens são ordenados por valor
    crescente e percorridos em ordem crescente de TAMANHO de combinação, então
    pra cada soma a combinação guardada é sempre a menor/mais enxuta
    encontrada primeiro (`mapa.setdefault` não sobrescreve).

    Se `teto_centavos` for informado, poda qualquer ramo cuja soma parcial já
    ultrapasse esse teto - usado aqui pra não perder tempo somando
    combinações de pagamentos maiores que qualquer soma de OBs que faça
    sentido consultar depois. `limite` é a mesma trava de `LIMITE_COMBINACOES`:
    se o número de combinações possíveis (n escolhe k) for grande demais pra
    um dado tamanho, esse tamanho é pulado."""
    ordenado = sorted(itens, key=lambda it: _centavos(valor_fn(it)))
    valores = [_centavos(valor_fn(it)) for it in ordenado]
    n = len(ordenado)
    mapa = {}

    for tamanho in range(1, min(max_tamanho, n) + 1):
        if math.comb(n, tamanho) > limite:
            continue
        escolhidos = [0] * tamanho

        def rec(inicio, k, soma_parcial):
            if k == tamanho:
                mapa.setdefault(soma_parcial, tuple(ordenado[i] for i in escolhidos))
                return
            limite_i = n - (tamanho - k)
            for i in range(inicio, limite_i + 1):
                nova_soma = soma_parcial + valores[i]
                if teto_centavos is not None and nova_soma > teto_centavos:
                    # itens[i:] só tem valores >= valores[i] (lista ordenada),
                    # então nenhum item a partir daqui cabe mais no teto
                    break
                escolhidos[k] = i
                rec(i + 1, k + 1, nova_soma)

        rec(0, 0, 0)

    return mapa


def _buscar_combo_obs_pagamentos(obs_disponiveis, pagamentos_disponiveis, max_tamanho_ob, max_tamanho_pag):
    """Última tentativa de conciliação (prioridade 4, ver docstring do
    módulo): quando uma OB isolada não bate com nenhum pagamento nem
    combinação de pagamentos, tenta somar 2 ou mais OBs entre si (dentre as
    que já sobraram sem match) e ver se essa soma bate com um pagamento ou
    combinação de pagamentos ainda disponível - caso real de várias remessas
    (OBs) que na prática saíram juntas num único lote/pagamento.

    Para não repetir uma busca combinatória cara a cada combinação de OBs
    testada, as somas possíveis do lado dos pagamentos são pré-computadas
    UMA VEZ (`_enumerar_somas`), e cada combinação de OBs só faz uma consulta
    O(1) nesse mapa em vez de refazer a busca inteira do zero. Isso troca
    "combinações de OBs" x "busca exaustiva de pagamentos" (multiplicativo,
    caro) por "combinações de OBs" + "combinações de pagamentos" (aditivo).

    Diferente da prioridade 1-3 (que prioriza pagamentos fisicamente
    próximos entre si via `_score_proximidade`), esta etapa não aplica esse
    desempate por proximidade - prioriza apenas o menor número de itens dos
    dois lados, o que já cobre o caso típico (poucas OBs somadas contra
    poucos pagamentos somados).

    `obs_disponiveis` é uma lista de (linha, codigo, valor); a busca
    considera combinações de tamanho `2` até `max_tamanho_ob` (limitado ao
    número de OBs disponíveis). Devolve (combo_obs, combo_pagamentos) ou
    None se nada bater. combo_obs é uma tupla de (linha, codigo, valor);
    combo_pagamentos é uma tupla de (linha, quantia)."""
    obs_ordenadas = sorted(obs_disponiveis, key=lambda x: x[2], reverse=True)
    n_obs = len(obs_ordenadas)
    max_tamanho_ob = min(max_tamanho_ob, n_obs)
    if max_tamanho_ob < 2 or not pagamentos_disponiveis:
        return None

    # teto: nenhuma combinação de OBs testada aqui passa da soma das
    # `max_tamanho_ob` maiores OBs disponíveis - usado pra podar a
    # enumeração de somas de pagamentos e não desperdiçar tempo somando
    # combinações de pagamentos maiores que qualquer alvo plausível
    teto_centavos = sum(_centavos(v) for _, _, v in obs_ordenadas[:max_tamanho_ob])

    mapa_pagamentos = _enumerar_somas(
        pagamentos_disponiveis, valor_fn=lambda p: p[1],
        max_tamanho=max_tamanho_pag, teto_centavos=teto_centavos,
    )
    if not mapa_pagamentos:
        return None

    for tamanho in range(2, max_tamanho_ob + 1):
        if math.comb(n_obs, tamanho) > LIMITE_COMBINACOES:
            # combinatória grande demais pra esse tamanho de combo de OBs -
            # pula pro próximo tamanho em vez de arriscar travar
            continue
        for combo_obs in itertools.combinations(obs_ordenadas, tamanho):
            alvo = sum(_centavos(valor) for _, _, valor in combo_obs)
            combo_pagamentos = mapa_pagamentos.get(alvo)
            if combo_pagamentos is not None:
                return combo_obs, combo_pagamentos

    return None


def conciliar_pagamentos_obs_aba(ws, max_combinacao=15, max_combinacao_obs=6):
    """Concilia pagamentos com OBs dentro de UMA aba. Devolve um dict com o
    relatório, ou None se a aba não tiver a estrutura esperada (bloco de
    Pagamentos e/ou tabela OB/VALOR)."""
    linha_titulo_pag = localizar_titulo(ws, "PAGAMENTOS")
    if linha_titulo_pag is None:
        return None

    col_ob, col_valor, linha_header_ob = localizar_bloco_ob(ws)
    if col_ob is None:
        return None

    pagamentos = _linhas_pagamentos(ws, linha_titulo_pag)
    obs = _linhas_ob_com_linha(ws, col_ob, col_valor, linha_header_ob + 1)

    # OBs maiores primeiro: tende a reduzir ambiguidade na hora de casar combinações
    obs_ordenadas = sorted(obs, key=lambda x: x[2], reverse=True)

    disponiveis = list(pagamentos)  # (linha, quantia) ainda não atrelados a nenhuma OB
    grupos = []

    # --- Prioridade 1: match exato de UM único pagamento, pra TODAS as OBs
    #     primeiro (sem ambiguidade, não depende de nenhuma combinação) ---
    obs_pendentes = []
    for ob_row, ob_codigo, valor in obs_ordenadas:
        alvo = _centavos(valor)
        pagamento_exato = next(
            (p for p in disponiveis if _centavos(p[1]) == alvo), None
        )
        if pagamento_exato is None:
            obs_pendentes.append((ob_row, ob_codigo, valor))
            continue
        grupos.append({
            "ob_rows": [ob_row],
            "ob_codigos": [ob_codigo],
            "valor": valor,
            "pagamento_rows": [pagamento_exato[0]],
        })
        disponiveis.remove(pagamento_exato)

    # --- Prioridade 2: combinações de 2+ OBs (dentre as que sobraram) vs
    #     combinações de pagamentos - ver comentário na docstring do módulo
    #     sobre por que isso roda ANTES da prioridade 3. Repete a busca até
    #     não achar mais nenhuma combinação, já que cada match muda o que
    #     ainda está disponível dos dois lados. ---
    while len(obs_pendentes) >= 2 and disponiveis:
        resultado_combo = _buscar_combo_obs_pagamentos(
            obs_pendentes, disponiveis, max_combinacao_obs, max_combinacao
        )
        if resultado_combo is None:
            break

        combo_obs, combo_pagamentos = resultado_combo
        ob_rows = [ob_row for ob_row, _, _ in combo_obs]
        ob_codigos = [ob_codigo for _, ob_codigo, _ in combo_obs]
        valor_total = sum(valor for _, _, valor in combo_obs)
        linhas_pagamento = [linha for linha, _ in combo_pagamentos]

        # tenta achar qual pagamento pertence a qual OB dentro do combo,
        # só pra decidir a cor de cada um na hora de pintar (ver
        # `_particionar_combo_por_ob`); se não achar uma divisão exata,
        # o grupo inteiro é pintado com uma cor só, como antes.
        particoes = _particionar_combo_por_ob(combo_obs, combo_pagamentos, max_combinacao)

        grupos.append({
            "ob_rows": ob_rows,
            "ob_codigos": ob_codigos,
            "valor": valor_total,
            "pagamento_rows": linhas_pagamento,
            "particoes_pintura": particoes,
        })

        usados_obs = set(ob_rows)
        obs_pendentes = [o for o in obs_pendentes if o[0] not in usados_obs]
        usados_pag = set(linhas_pagamento)
        disponiveis = [p for p in disponiveis if p[0] not in usados_pag]

    # --- Prioridades 3 e 4: pra quem ainda sobrou, OB isolada vs combinação
    #     de 2+ pagamentos (`_buscar_subconjunto` já cuida internamente de
    #     priorizar pagamentos próximos entre si antes de aceitar qualquer
    #     combinação exaustiva) ---
    obs_sem_match = []
    for ob_row, ob_codigo, valor in obs_pendentes:
        alvo = _centavos(valor)
        combo = _buscar_subconjunto(disponiveis, alvo, max_combinacao)
        if combo is None:
            obs_sem_match.append((ob_row, ob_codigo, valor))
            continue
        linhas_pagamento = [linha for linha, _ in combo]
        grupos.append({
            "ob_rows": [ob_row],
            "ob_codigos": [ob_codigo],
            "valor": valor,
            "pagamento_rows": linhas_pagamento,
        })
        usados = set(linhas_pagamento)
        disponiveis = [p for p in disponiveis if p[0] not in usados]

    # pinta os grupos conciliados. Cada "unidade visual" leva sua própria
    # cor: pra grupo de OB única (prioridades 1, 3 e 4) a unidade é o
    # grupo inteiro; pra grupo de OBs combinadas (prioridade 2) em que
    # achamos qual pagamento pertence a qual OB (`particoes_pintura`), cada
    # OB + seus pagamentos correspondentes vira uma unidade própria, com
    # sua própria cor - só cai numa cor só pro grupo inteiro se essa
    # divisão não foi possível. A fonte é sempre resetada pro padrão, pra
    # não deixar resíduo de cor de letra de alguma rodada anterior.
    cor_idx = 0
    for grupo in grupos:
        particoes = grupo.get("particoes_pintura")
        if particoes:
            unidades = [
                {"ob_rows": [p["ob_row"]], "pagamento_rows": p["pagamento_rows"]}
                for p in particoes
            ]
        else:
            unidades = [{"ob_rows": grupo["ob_rows"], "pagamento_rows": grupo["pagamento_rows"]}]

        for unidade in unidades:
            cor = CORES_CONCILIACAO[cor_idx % len(CORES_CONCILIACAO)]
            cor_idx += 1
            fill = PatternFill("solid", fgColor=cor)

            for ob_row in unidade["ob_rows"]:
                for col in (col_ob, col_valor):
                    cell = ws.cell(row=ob_row, column=col)
                    cell.fill = fill
                    cell.font = NORMAL_FONT

            for linha in unidade["pagamento_rows"]:
                for col in range(1, NUM_COLS_PRINCIPAL + 1):
                    cell = ws.cell(row=linha, column=col)
                    cell.fill = fill
                    cell.font = NORMAL_FONT

    return {
        "aba": ws.title,
        "grupos": grupos,
        "obs_sem_match": obs_sem_match,
        "pagamentos_sem_ob": disponiveis,
        "pagamentos_todos": pagamentos,
        "total_obs": len(obs),
    }


def conciliar_pagamentos_obs(caminho, max_combinacao=6, max_combinacao_obs=6, abas=None):
    wb = openpyxl.load_workbook(caminho)
    abas_selecionadas = _normalizar_abas(abas)
    abas_alvo = abas_selecionadas or wb.sheetnames

    resultados = []
    for nome in abas_alvo:
        if nome not in wb.sheetnames:
            print(f"Aba '{nome}' não encontrada. Ignorando.")
            continue

        ws = wb[nome]
        resultado = conciliar_pagamentos_obs_aba(
            ws, max_combinacao=max_combinacao, max_combinacao_obs=max_combinacao_obs
        )
        if resultado is not None:
            resultados.append(resultado)

    wb.save(caminho)

    if not resultados:
        print("Nenhuma aba com bloco de Pagamentos + tabela OB/VALOR foi encontrada.")
        return

    for res in resultados:
        n_obs_conciliadas = sum(len(grupo["ob_rows"]) for grupo in res["grupos"])
        print(f"\n[{res['aba']}] OBs conciliadas: {n_obs_conciliadas} de {res['total_obs']} (em {len(res['grupos'])} grupo(s))")
        for grupo in res["grupos"]:
            n_pag = len(grupo["pagamento_rows"])
            composicao = "1 pagamento" if n_pag == 1 else f"{n_pag} pagamentos somados"
            obs_str = " + ".join(str(c) for c in grupo["ob_codigos"])
            print(f"  OB(s) {obs_str} (R$ {grupo['valor']:.2f} total) <- {composicao}, linha(s) {grupo['pagamento_rows']}")

        if res["obs_sem_match"]:
            print(f"  ⚠️  ATENÇÃO: {len(res['obs_sem_match'])} OB(s) SEM correspondência encontrada nesta aba!")
            for ob_row, ob_codigo, valor in res["obs_sem_match"]:
                print(f"    OB {ob_codigo} (R$ {valor:.2f})")

        if res["pagamentos_sem_ob"]:
            print(f"  Pagamentos sem OB correspondente: {len(res['pagamentos_sem_ob'])}")
            for linha, quantia in res["pagamentos_sem_ob"]:
                print(f"    Linha {linha} (R$ {quantia:.2f})")

    # Resumo global: como toda OB deveria, em tese, ter um pagamento (ou
    # combinação de pagamentos) correspondente, qualquer OB sem match é um
    # sinal de que algo precisa ser conferido manualmente. Por isso esse
    # aviso fica bem destacado no final, reunindo todas as abas, e não só
    # espalhado aba por aba.
    total_obs_sem_match = sum(len(res["obs_sem_match"]) for res in resultados)
    if total_obs_sem_match > 0:
        print("\n" + "=" * 60)
        print(f"⚠️  ATENÇÃO: {total_obs_sem_match} OB(s) NÃO conciliada(s) no total!")
        print("=" * 60)
        for res in resultados:
            if res["obs_sem_match"]:
                for ob_row, ob_codigo, valor in res["obs_sem_match"]:
                    print(f"  [{res['aba']}] OB {ob_codigo} (R$ {valor:.2f}) - linha {ob_row}")
        print("=" * 60)
    else:
        print("\n✅ Todas as OBs foram conciliadas com sucesso em todas as abas.")
