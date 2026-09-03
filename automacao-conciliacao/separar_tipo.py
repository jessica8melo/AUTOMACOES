"""
Passo 2 (chamado por test.py, via `processar_aba`, uma aba por vez): separa
os lançamentos em Recebimentos e Pagamentos dentro da mesma aba, empilhando 
Recebimentos no topo e Pagamentos embaixo.

Regra usada: qualquer valor da coluna "Tipo" que COMEÇA com "Recebimento"
vai para o bloco de Recebimentos; qualquer valor que COMEÇA com "Pagamento"
vai para o bloco de Pagamentos. Cobre variações como "Recebimento Div." /
"Pagamento Div.".

A aba só é alterada se tiver a estrutura esperada (cabeçalho com "Tipo" na
coluna B); caso contrário `processar_aba` devolve None e é ignorada por
quem chamou.
"""

import datetime
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FONT_NAME = "Arial"
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
BLOCO_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
BLOCO_FILL_RECEB = PatternFill("solid", fgColor="2E7D32")
BLOCO_FILL_PAG = PatternFill("solid", fgColor="B71C1C")
BLOCO_FILL_OUTROS = PatternFill("solid", fgColor="616161")
NORMAL_FONT = Font(name=FONT_NAME, size=10)

THIN = Side(style="thin", color="B7B7B7")
THICK = Side(style="thick", color="000000")

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

NUM_COLS = 7  # A:G -> Linha, Tipo, Código, Data Transação, Quantia, Status, Justificativas
COL_QUANTIA = 5
COL_DATA = 4


def classificar(tipo) -> str:
    """Retorna 'Recebimento', 'Pagamento' ou 'Outro' com base no texto do Tipo."""
    if not tipo:
        return "Outro"
    t = str(tipo).strip().lower()
    if t.startswith("recebimento"):
        return "Recebimento"
    if t.startswith("pagamento"):
        return "Pagamento"
    return "Outro"


def ultima_linha_com_dados(ws, max_col):
    """Última linha (>=1) que tem algum valor não vazio em alguma das
    primeiras `max_col` colunas. Necessário porque a aba pode ter
    formatação/dimensão até a linha 1000 sem dado real."""
    last = 1
    for r, row in enumerate(
        ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=max_col, values_only=True),
        start=2,
    ):
        if any(v is not None and v != "" for v in row):
            last = r
    return last


def localizar_col_ob(ws, ultima_linha, max_col):
    """Procura no cabeçalho (linha 1) uma célula 'OB' seguida, na coluna
    seguinte, por uma célula 'VALOR'. Retorna (col_ob, col_valor) ou
    (None, None). Em seguida, estende o início do bloco para a esquerda
    enquanto encontrar colunas com dados contíguas antes de 'OB' (ex.: uma
    coluna sem cabeçalho tipo 'OK' colada ao lado do OB), sem nunca invadir
    a tabela principal (colunas 1..NUM_COLS)."""
    col_ob = None
    for c in range(1, max_col + 1):
        v = ws.cell(row=1, column=c).value
        if v and str(v).strip().upper() == "OB":
            col_ob = c
            break
    if col_ob is None:
        return None, None, None

    col_valor = col_ob + 1
    v_valor = ws.cell(row=1, column=col_valor).value
    if not v_valor or str(v_valor).strip().upper() != "VALOR":
        return None, None, None

    extra_inicio = col_ob
    col = col_ob - 1
    while col > NUM_COLS:
        tem_dado = any(
            ws.cell(row=r, column=col).value not in (None, "")
            for r in range(2, ultima_linha + 1)
        )
        if tem_dado:
            extra_inicio = col
            col -= 1
        else:
            break

    return col_ob, col_valor, extra_inicio


def coletar_conteudo_fora_do_padrao(ws, ultima_linha, max_col, extra_inicio, col_valor):
    """Lista células com conteúdo fora das colunas conhecidas (tabela
    principal 1..NUM_COLS e bloco extra extra_inicio..col_valor), para
    avisar o usuário em vez de apagar silenciosamente."""
    avisos = []
    faixas_conhecidas = set(range(1, NUM_COLS + 1))
    if extra_inicio is not None:
        faixas_conhecidas |= set(range(extra_inicio, col_valor + 1))
    for r in range(2, ultima_linha + 1):
        for c in range(1, max_col + 1):
            if c in faixas_conhecidas:
                continue
            v = ws.cell(row=r, column=c).value
            if v not in (None, ""):
                avisos.append((r, get_column_letter(c), v))
    return avisos


def escrever_cabecalho_colunas(ws, row, headers, col_inicial=1):
    for off, titulo in enumerate(headers):
        c = ws.cell(row=row, column=col_inicial + off, value=titulo)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def escrever_titulo_bloco(ws, row, texto, fill, num_cols, col_inicial=1):
    ws.merge_cells(start_row=row, start_column=col_inicial, end_row=row,
                    end_column=col_inicial + num_cols - 1)
    c = ws.cell(row=row, column=col_inicial, value=texto)
    c.font = BLOCO_FONT
    c.fill = fill
    c.alignment = CENTER
    for col in range(col_inicial, col_inicial + num_cols):
        ws.cell(row=row, column=col).fill = fill


def escrever_linhas(ws, start_row, linhas, col_valor_idx=None, col_data_idx=None, col_inicial=1):
    """Escreve as linhas de dados a partir de start_row. Retorna a última linha usada
    (start_row - 1 se não houver linhas)."""
    r = start_row
    for linha in linhas:
        for c_off, valor in enumerate(linha, start=0):
            c_idx = c_off + 1  # posição dentro do bloco (1-based)
            cell = ws.cell(row=r, column=col_inicial + c_off, value=valor)
            cell.font = NORMAL_FONT
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            if col_valor_idx is not None and c_idx == col_valor_idx:
                cell.number_format = "#,##0.00"
                cell.alignment = RIGHT
            elif col_data_idx is not None and c_idx == col_data_idx and isinstance(valor, (datetime.datetime, datetime.date)):
                cell.number_format = "d-mmm-yy"
                cell.alignment = CENTER
            elif c_idx == 1:
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT
        r += 1
    return r - 1


def aplicar_borda_grossa_inferior(ws, row, col_inicial, num_cols):
    for col in range(col_inicial, col_inicial + num_cols):
        cell = ws.cell(row=row, column=col)
        existing = cell.border
        cell.border = Border(left=existing.left, right=existing.right,
                              top=existing.top, bottom=THICK)


def limpar_aba(ws):
    """Remove valores, formatação e merges existentes, mantendo a aba (mesmo
    nome/posição) para reconstruir o conteúdo do zero dentro dela."""
    for mc in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mc))
    max_row = ws.max_row
    max_col = ws.max_column
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.value = None
            cell.font = Font(name=FONT_NAME, size=10)
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border()
            cell.alignment = Alignment()
    ws.auto_filter.ref = None


def corrigir_data_serial(valor, epoch):
    """Converte a Data Transação para datetime, usando o
    epoch da própria planilha. Valores que já são datetime, ou que 
    não parecem data, são devolvidos sem alteração."""
    if isinstance(valor, (datetime.datetime, datetime.date)):
        return valor
    if isinstance(valor, bool):
        return valor
    if isinstance(valor, (int, float)):
        try:
            return from_excel(valor, epoch)
        except Exception:
            return valor
    return valor


def processar_aba(ws, epoch):
    """Processa uma aba. Retorna um dict com o relatório, ou None se a aba
    não tiver a estrutura esperada (e nesse caso não é alterada)."""
    cabecalho_b = ws.cell(row=1, column=2).value
    if not cabecalho_b or "tipo" not in str(cabecalho_b).strip().lower():
        return None

    max_col = max(ws.max_column, NUM_COLS + 2)
    ultima_linha = ultima_linha_com_dados(ws, max_col)

    headers = [ws.cell(row=1, column=c).value for c in range(1, NUM_COLS + 1)]

    recebimentos, pagamentos, outros = [], [], []
    for r in range(2, ultima_linha + 1):
        linha_valores = [ws.cell(row=r, column=c).value for c in range(1, NUM_COLS + 1)]
        if linha_valores[0] is None:
            continue
        linha_valores[COL_DATA - 1] = corrigir_data_serial(linha_valores[COL_DATA - 1], epoch)
        categoria = classificar(linha_valores[1])
        if categoria == "Recebimento":
            recebimentos.append(linha_valores)
        elif categoria == "Pagamento":
            pagamentos.append(linha_valores)
        else:
            outros.append(linha_valores)

    col_ob, col_valor, extra_inicio = localizar_col_ob(ws, ultima_linha, max_col)

    headers_extra = None
    linhas_extra = []
    if col_ob is not None:
        headers_extra = [ws.cell(row=1, column=c).value for c in range(extra_inicio, col_valor + 1)]
        for r in range(2, ultima_linha + 1):
            valores = [ws.cell(row=r, column=c).value for c in range(extra_inicio, col_valor + 1)]
            if any(v is not None and v != "" for v in valores):
                linhas_extra.append(valores)

    avisos = coletar_conteudo_fora_do_padrao(ws, ultima_linha, max_col, extra_inicio, col_valor)

    # --- Reconstrói a aba ---
    limpar_aba(ws)
    linha_atual = 1

    escrever_titulo_bloco(ws, linha_atual, f"RECEBIMENTOS ({len(recebimentos)})",
                           BLOCO_FILL_RECEB, NUM_COLS)
    linha_atual += 1
    escrever_cabecalho_colunas(ws, linha_atual, headers)
    linha_header_receb = linha_atual
    linha_atual += 1
    ultima_linha_receb = escrever_linhas(ws, linha_atual, recebimentos,
                                          col_valor_idx=COL_QUANTIA, col_data_idx=COL_DATA)
    if not recebimentos:
        ultima_linha_receb = linha_atual - 1

    linha_divisoria = max(ultima_linha_receb, linha_header_receb)
    aplicar_borda_grossa_inferior(ws, linha_divisoria, 1, NUM_COLS)

    linha_atual = linha_divisoria + 2

    linha_titulo_pag = linha_atual
    escrever_titulo_bloco(ws, linha_atual, f"PAGAMENTOS ({len(pagamentos)})",
                           BLOCO_FILL_PAG, NUM_COLS)
    linha_atual += 1
    escrever_cabecalho_colunas(ws, linha_atual, headers)
    linha_atual += 1
    inicio_dados_pag = linha_atual
    ultima_linha_pag = escrever_linhas(ws, linha_atual, pagamentos,
                                        col_valor_idx=COL_QUANTIA, col_data_idx=COL_DATA)
    if not pagamentos:
        ultima_linha_pag = linha_atual - 1

    if headers_extra is not None:
        col_inicial_extra = NUM_COLS + 2  # deixa uma coluna de espaço
        num_cols_extra = len(headers_extra)

        ws.merge_cells(start_row=linha_titulo_pag, start_column=col_inicial_extra,
                        end_row=linha_titulo_pag, end_column=col_inicial_extra + num_cols_extra - 1)
        c = ws.cell(row=linha_titulo_pag, column=col_inicial_extra, value="OB / VALOR")
        c.font = BLOCO_FONT
        c.fill = BLOCO_FILL_PAG
        c.alignment = CENTER
        for off in range(num_cols_extra):
            ws.cell(row=linha_titulo_pag, column=col_inicial_extra + off).fill = BLOCO_FILL_PAG

        escrever_cabecalho_colunas(ws, linha_titulo_pag + 1, headers_extra, col_inicial=col_inicial_extra)

        r = inicio_dados_pag
        for valores in linhas_extra:
            for off, v in enumerate(valores):
                cell = ws.cell(row=r, column=col_inicial_extra + off, value=v)
                cell.font = NORMAL_FONT
                cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
                if off == num_cols_extra - 1:  # última coluna do bloco = VALOR
                    cell.number_format = "#,##0.00"
                    cell.alignment = RIGHT
                else:
                    cell.alignment = CENTER
            r += 1

        for off in range(num_cols_extra):
            letra = ws.cell(row=1, column=col_inicial_extra + off).column_letter
            ws.column_dimensions[letra].width = 14

    if outros:
        linha_atual = ultima_linha_pag + 2
        escrever_titulo_bloco(ws, linha_atual, f"OUTROS ({len(outros)})",
                               BLOCO_FILL_OUTROS, NUM_COLS)
        linha_atual += 1
        escrever_cabecalho_colunas(ws, linha_atual, headers)
        linha_atual += 1
        escrever_linhas(ws, linha_atual, outros, col_valor_idx=COL_QUANTIA, col_data_idx=COL_DATA)

    larguras = {"A": 7, "B": 18, "C": 11, "D": 15, "E": 13, "F": 18, "G": 32}
    for col, w in larguras.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    return {
        "aba": ws.title,
        "recebimentos": len(recebimentos),
        "pagamentos": len(pagamentos),
        "outros": len(outros),
        "extra": len(linhas_extra) if headers_extra is not None else None,
        "avisos": avisos,
    }