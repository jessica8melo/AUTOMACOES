"""
Passo 2: em cada aba, concilia pagamento(s) com OB dentro dessa mesma aba.
A busca por match segue uma ordem de prioridade:

    1) Match exato de UM único pagamento cujo valor bate com o de UMA OB.
       Feito primeiro, pra TODAS as OBs, sem nenhuma ambiguidade (não
       depende de olhar combinações).

    2) Para as OBs que ainda sobrarem, combinações de 2 ou mais pagamentos
       cuja soma bate com o valor de UMA OB isolada, dando preferência para
       pagamentos que estejam fisicamente próximos entre si na planilha
       (idealmente na linha logo abaixo ou logo acima um do outro).

    3) Se nada acima servir, vale qualquer combinação exata encontrada pela
       busca exaustiva (com poda), mesmo que os pagamentos estejam
       espalhados/longe uns dos outros.

Cada grupo de match é sempre UMA única OB (nunca 2+ OBs somadas contra um
mesmo combo de pagamentos). Quando acha, pinta a(s) linha(s) do(s)
pagamento(s) e a linha da OB com a mesma cor, deixando visualmente claro
quem está atrelado a quem.

Abas sem bloco de Pagamentos e/ou sem tabela OB/VALOR são ignoradas (nada é
alterado nelas).
"""

import math

from openpyxl.styles import PatternFill

from processar_pagamentos import (
    CORES_CONCILIACAO, NORMAL_FONT, NUM_COLS_PRINCIPAL,
    localizar_titulo, localizar_bloco_ob, _centavos,
)

# trava de segurança: para um dado tamanho de combinação, se o número de
# combinações possíveis (n escolhe k) passar disso, esse tamanho é pulado em
# vez de arriscar travar o script numa aba com muitos itens sem match
LIMITE_COMBINACOES = 200_000


def _linhas_pagamentos(ws, linha_titulo_pag):
    """Lê a tabela principal de Pagamentos (colunas A:G) dessa aba,
    começando logo após o título 'PAGAMENTOS (N)' e o cabeçalho de colunas
    que vem em seguida. Devolve [(linha, quantia)] enquanto a coluna
    'Linha' (A) tiver valor."""
    linhas = []
    r = linha_titulo_pag + 2  # +1 = cabeçalho de colunas, +2 = primeira linha de dado
    while ws.cell(row=r, column=1).value is not None:
        quantia = ws.cell(row=r, column=5).value  # coluna E = Quantia
        if quantia is not None:
            linhas.append((r, quantia))
        r += 1
    return linhas


def _linhas_ob_com_linha(ws, col_ob, col_valor, linha_inicial):
    """Como _ler_linhas_ob, mas guardando também o número da linha (usado
    pra pintar depois). Também para de ler se a coluna VALOR trouxer algo
    que não seja número (ver comentário em _ler_linhas_ob)."""
    linhas = []
    r = linha_inicial
    while True:
        v_ob = ws.cell(row=r, column=col_ob).value
        v_valor = ws.cell(row=r, column=col_valor).value
        if v_ob is None and v_valor is None:
            break
        if v_valor is not None and not isinstance(v_valor, (int, float)):
            break
        if v_valor is not None:
            linhas.append((r, v_ob, v_valor))
        r += 1
    return linhas


def _score_proximidade(combo):
    """Gera uma penalidade para combinações de 2+ pagamentos cujas linhas
    estão mais distantes entre si na planilha. Menor score é melhor. Dá
    peso extra para sequências contíguas (linhas vizinhas), que é o que
    queremos priorizar quando um único pagamento não bate exato com a OB.

    Usada apenas para desempate ENTRE combinações de 2+ itens - o caso de
    um único pagamento com valor exato é tratado à parte, antes disso, em
    `_buscar_subconjunto`."""
    linhas = sorted(linha for linha, _ in combo)
    span = linhas[-1] - linhas[0]
    gaps = [b - a for a, b in zip(linhas, linhas[1:])]
    contiguas = sum(1 for gap in gaps if gap == 1)
    return (span, sum(gaps), -contiguas, linhas[0])


def _buscar_subconjunto(pagamentos_disponiveis, alvo_centavos, max_tamanho):
    """Procura pagamento(s) cuja soma bate exatamente com `alvo_centavos`,
    seguindo a ordem de prioridade.

    Devolve a combinação (tupla de (linha, quantia)) ou None se nada bater.

    A busca em si usa poda (branch and bound) em vez de testar todas as
    combinações via itertools.combinations: os pagamentos são ordenados por
    valor, e a recursão corta um ramo assim que a soma parcial já passa do
    alvo (como a lista está em ordem crescente, os itens seguintes só
    aumentam a soma, então não há por que continuar). Sem isso, uma aba com
    ~50-80 pagamentos e várias OBs sem correspondência podia levar minutos
    (ou até travar na prática) tentando C(n, 6) combinações repetidamente."""
    itens = sorted(pagamentos_disponiveis, key=lambda x: _centavos(x[1]))
    valores = [_centavos(q) for _, q in itens]
    n = len(itens)
    max_tamanho = min(max_tamanho, n)

    # --- Prioridade 1: pagamento único com valor exatamente igual ao da OB ---
    for item, valor_centavos in zip(itens, valores):
        if valor_centavos == alvo_centavos:
            return (item,)

    # --- Prioridades 2 e 3: combinações de 2+ itens, do menor pro maior
    #     tamanho, com desempate por proximidade entre linhas ---
    melhor_combo = None
    melhor_score = None

    for tamanho in range(2, max_tamanho + 1):
        if math.comb(n, tamanho) > LIMITE_COMBINACOES:
            # combinatória grande demais pra esse tamanho - pula pro
            # próximo em vez de arriscar travar
            continue
        combo = _buscar_tamanho_fixo(itens, valores, alvo_centavos, tamanho)
        if combo is None:
            continue

        score = _score_proximidade(combo)
        if melhor_combo is None or score < melhor_score:
            melhor_combo = combo
            melhor_score = score

    return melhor_combo


def _buscar_tamanho_fixo(itens, valores, alvo, tamanho):
    """Busca com poda por uma combinação de exatamente `tamanho` itens (já
    ordenados por valor crescente em `itens`/`valores`) cuja soma bata com
    `alvo`. Devolve a tupla de itens ou None. Quando houver múltiplas
    combinações exatas desse mesmo tamanho, a melhor é escolhida pelo
    critério de proximidade entre linhas (`_score_proximidade`)."""
    n = len(itens)
    escolhidos = [0] * tamanho
    melhor_combo = None

    def rec(inicio, k, soma_parcial):
        nonlocal melhor_combo
        if k == tamanho:
            if soma_parcial == alvo:
                combo = tuple(itens[i] for i in escolhidos)
                if melhor_combo is None:
                    melhor_combo = combo
                else:
                    atual_score = _score_proximidade(combo)
                    melhor_score = _score_proximidade(melhor_combo)
                    if atual_score < melhor_score:
                        melhor_combo = combo
            return
        # não vale a pena tentar índices onde não sobram itens suficientes
        # para completar os `tamanho - k` restantes
        limite = n - (tamanho - k)
        for i in range(inicio, limite + 1):
            nova_soma = soma_parcial + valores[i]
            if nova_soma > alvo:
                # itens[i:] só tem valores >= valores[i] (lista ordenada),
                # então nenhum item a partir daqui pode mais servir
                break
            escolhidos[k] = i
            rec(i + 1, k + 1, nova_soma)
        return None

    rec(0, 0, 0)
    return melhor_combo


def _contar_pagamentos_pendentes(ws):
    """Conta quantos pagamentos sobram sem OB depois SÓ da Prioridade 1
    (match exato de um único pagamento) - ou seja, o tamanho do "pool" de
    pagamentos que as Prioridades 2-4 (combinações) vão ter que vasculhar
    numa aba. Esse número não depende de nenhum teto (a Prioridade 1 não
    usa `max_combinacao`), então dá pra calcular baratinho, sem rodar a
    busca de combinações nenhuma vez - é usado por
    `conciliar_pagamentos_obs_aba` pra decidir o teto INICIAL da
    escalonada (ver docstring lá).

    Devolve a contagem, ou None se a aba não tiver a estrutura esperada
    (mesmo critério de `_conciliar_pagamentos_obs_aba_com_teto`: precisa
    ter bloco de Pagamentos e tabela OB/VALOR)."""
    linha_titulo_pag = localizar_titulo(ws, "PAGAMENTOS")
    if linha_titulo_pag is None:
        return None

    col_ob, col_valor, linha_header_ob = localizar_bloco_ob(ws)
    if col_ob is None:
        return None

    pagamentos = _linhas_pagamentos(ws, linha_titulo_pag)
    obs = _linhas_ob_com_linha(ws, col_ob, col_valor, linha_header_ob + 1)

    disponiveis = list(pagamentos)
    for _, _, valor in obs:
        alvo = _centavos(valor)
        pagamento_exato = next(
            (p for p in disponiveis if _centavos(p[1]) == alvo), None
        )
        if pagamento_exato is not None:
            disponiveis.remove(pagamento_exato)

    return len(disponiveis)


def _conciliar_nucleo(pagamentos_disponiveis, obs_disponiveis, max_combinacao):
    """Roda as Prioridades 1-3 (ver docstring do módulo) sobre as LISTAS
    dadas - não lê nem escreve na planilha, só calcula. Usada tanto pela
    tentativa única quanto por cada rodada do escalonamento cumulativo em
    `conciliar_pagamentos_obs_aba` (por isso não sabe nada sobre teto
    escalonado nem sobre pintura - isso é responsabilidade de quem chama).

    Devolve (grupos, obs_sem_match, pagamentos_sobrando):
        - grupos: lista de dicts com "ob_rows"/"ob_codigos"/"valor"/
          "pagamento_rows" (sempre uma única OB por grupo).
        - obs_sem_match: [(ob_row, ob_codigo, valor), ...] que não acharam
          par usando os itens disponíveis.
        - pagamentos_sobrando: [(linha, quantia), ...] que sobraram sem
          nenhuma OB correspondente."""
    # OBs maiores primeiro: tende a reduzir ambiguidade na hora de casar combinações
    obs_ordenadas = sorted(obs_disponiveis, key=lambda x: x[2], reverse=True)

    disponiveis = list(pagamentos_disponiveis)  # (linha, quantia) ainda não atrelados a nenhuma OB
    grupos = []

    # --- Prioridade 1: match exato de UM único pagamento, pra TODAS as OBs
    #     primeiro (sem ambiguidade, não depende de nenhuma combinação) ---
    obs_pendentes = []
    for ob_row, ob_codigo, valor in obs_ordenadas:
        alvo = _centavos(valor)
        pagamento_exato = next(
            (p for p in disponiveis if _centavos(p[1]) == alvo), None
        )
        if pagamento_exato is None:
            obs_pendentes.append((ob_row, ob_codigo, valor))
            continue
        grupos.append({
            "ob_rows": [ob_row],
            "ob_codigos": [ob_codigo],
            "valor": valor,
            "pagamento_rows": [pagamento_exato[0]],
        })
        disponiveis.remove(pagamento_exato)

    # --- Prioridades 2 e 3: pra quem ainda sobrou, OB isolada vs combinação
    #     de 2+ pagamentos (`_buscar_subconjunto` já cuida internamente de
    #     priorizar pagamentos próximos entre si antes de aceitar qualquer
    #     combinação exaustiva) ---
    obs_sem_match = []
    for ob_row, ob_codigo, valor in obs_pendentes:
        alvo = _centavos(valor)
        combo = _buscar_subconjunto(disponiveis, alvo, max_combinacao)
        if combo is None:
            obs_sem_match.append((ob_row, ob_codigo, valor))
            continue
        linhas_pagamento = [linha for linha, _ in combo]
        grupos.append({
            "ob_rows": [ob_row],
            "ob_codigos": [ob_codigo],
            "valor": valor,
            "pagamento_rows": linhas_pagamento,
        })
        usados = set(linhas_pagamento)
        disponiveis = [p for p in disponiveis if p[0] not in usados]

    return grupos, obs_sem_match, disponiveis


def _pintar_grupos(ws, grupos, col_ob, col_valor):
    """Pinta na planilha todos os `grupos` já conciliados (formato de
    `_conciliar_nucleo`). Extraída à parte pra poder rodar UMA vez só, no
    final, sobre os grupos ACUMULADOS de todas as rodadas do escalonamento
    em `conciliar_pagamentos_obs_aba` - assim a numeração de cores é
    contínua e nada fica pintado por uma rodada e sobrescrito por outra.

    Cada grupo é sempre uma única OB e vira sua própria unidade visual,
    com sua própria cor. A fonte é sempre resetada pro padrão, pra não
    deixar resíduo de cor de letra de alguma rodada anterior."""
    cor_idx = 0
    for grupo in grupos:
        cor = CORES_CONCILIACAO[cor_idx % len(CORES_CONCILIACAO)]
        cor_idx += 1
        fill = PatternFill("solid", fgColor=cor)

        for ob_row in grupo["ob_rows"]:
            for col in (col_ob, col_valor):
                cell = ws.cell(row=ob_row, column=col)
                cell.fill = fill
                cell.font = NORMAL_FONT

        for linha in grupo["pagamento_rows"]:
            for col in range(1, NUM_COLS_PRINCIPAL + 1):
                cell = ws.cell(row=linha, column=col)
                cell.fill = fill
                cell.font = NORMAL_FONT


def conciliar_pagamentos_obs_aba(ws, max_combinacao=None, max_combinacao_minimo=6):
    """Concilia pagamentos com OBs dentro de UMA aba, de forma ESCALONADA e
    CUMULATIVA: tenta primeiro com um teto MÁXIMO de tamanho de combinação
    de pagamentos - a tentativa mais permissiva, que já cobre de cara
    qualquer combinação que uma tentativa menor também acharia.

    Qualquer OB (e os pagamentos que ela usou) que ache match nessa
    tentativa fica CONFIRMADA e sai do jogo - as rodadas seguintes, com
    teto menor, só voltam a mexer nas OBs que AINDA sobraram sem par,
    testando-as contra o que também sobrou de pagamentos. Ou seja,
    diferente de antes, um match achado com um teto grande não é jogado
    fora só porque outra OB (sem nenhuma relação com esse match) ainda não
    achou par e obriga o teto a continuar descendo.

    O teto vai sendo DIMINUÍDO (de 1 em 1) até no mínimo
    `max_combinacao_minimo`: um teto menor é mais conservador e pode
    resolver casos que um teto maior não resolveu - com valores repetidos
    (ex.: vários pagamentos de R$300,00 iguais), uma OB processada com
    teto alto pode "roubar" gananciosamente uma combinação grande de
    pagamentos que na verdade pertenceria a outra OB (ver comentário na
    Prioridade 2, acima); um teto menor evita essa combinação grande e
    pode deixar sobrar exatamente os pagamentos que a outra OB precisava.
    Isso continua valendo por rodada (dentro de uma mesma tentativa de
    teto, todas as OBs ainda pendentes são processadas juntas, na mesma
    ordem de prioridade de sempre) - só a forma de ACUMULAR entre tetos
    diferentes é que mudou.

    O teto MÁXIMO inicial é calculado por aba, como a quantidade de
    pagamentos que sobram sem OB depois da Prioridade 1 (match exato) -
    ver `_contar_pagamentos_pendentes`. Ou seja, o teto já começa grande o
    bastante pra cobrir QUALQUER combinação possível com os pagamentos
    daquela aba, e desce a partir daí até `max_combinacao_minimo`. Ex.: se
    sobraram 32 pagamentos sem match direto numa aba, a escalonada tenta
    teto 32, 31, 30, ... até `max_combinacao_minimo` (parando antes se
    algum teto já conciliar tudo que restava).

    NOTA sobre performance: como o pool de OBs e pagamentos pendentes só
    ENCOLHE a cada rodada (nunca é resetado), tamanhos de combinação que
    a trava `LIMITE_COMBINACOES` recusava numa rodada com pool cheio podem
    passar a ser viáveis nas rodadas seguintes, já com o pool reduzido -
    isso é uma vantagem adicional da versão cumulativa: além de não perder
    matches já achados, ela também aumenta a chance de matches que exigem
    combinações grandes (ex.: 8-9 pagamentos) serem encontrados, já que o
    pool relevante pra essa checagem vai diminuindo a cada match confirmado.

    `max_combinacao`, se informado, funciona como um teto ABSOLUTO por
    cima desse valor calculado - útil como trava de segurança em abas com
    MUITOS pagamentos sem match, pra não deixar a busca cara demais. Se
    omitido (None, padrão), não há teto artificial: o teto inicial é
    sempre o número de pagamentos pendentes daquela aba.

    Assim que não sobrar nenhuma OB sem match, para ali - não continua
    diminuindo o teto à toa.

    Se `max_combinacao_minimo >= teto inicial`, roda só uma vez com esse
    teto (sem escalonamento).

    A pintura das células só acontece UMA vez, no final, sobre todos os
    grupos acumulados de todas as rodadas (ver `_pintar_grupos`).

    Devolve um dict com o relatório consolidado, com uma chave extra
    "max_combinacao_usado" indicando o MENOR teto que chegou a ser
    tentado (só informativo - tetos diferentes podem ter resolvido OBs
    diferentes). Ou None se a aba não tiver a estrutura esperada (bloco de
    Pagamentos e/ou tabela OB/VALOR)."""
    linha_titulo_pag = localizar_titulo(ws, "PAGAMENTOS")
    if linha_titulo_pag is None:
        return None

    col_ob, col_valor, linha_header_ob = localizar_bloco_ob(ws)
    if col_ob is None:
        return None

    pagamentos = _linhas_pagamentos(ws, linha_titulo_pag)
    obs = _linhas_ob_com_linha(ws, col_ob, col_valor, linha_header_ob + 1)

    n_pendentes = _contar_pagamentos_pendentes(ws)
    if n_pendentes is None:
        return None

    teto_inicial = n_pendentes if max_combinacao is None else min(max_combinacao, n_pendentes)
    teto_inicial = max(teto_inicial, 1)  # nunca roda com teto menor que 1
    teto_minimo = min(max_combinacao_minimo, teto_inicial)

    obs_restantes = list(obs)
    pagamentos_restantes = list(pagamentos)
    grupos_totais = []
    teto_usado = teto_inicial

    for teto in range(teto_inicial, teto_minimo - 1, -1):
        if not obs_restantes:
            break

        teto_usado = teto
        grupos, obs_sem_match, pagamentos_restantes = _conciliar_nucleo(
            pagamentos_restantes, obs_restantes, max_combinacao=teto
        )

        n_resolvidas = len(obs_restantes) - len(obs_sem_match)
        grupos_totais.extend(grupos)
        obs_restantes = obs_sem_match

        if not obs_restantes:
            break
        if teto > teto_minimo:
            if n_resolvidas > 0:
                print(f"  ✅ [{ws.title}] {n_resolvidas} OB(s) conciliada(s) com teto {teto} (match(es) mantido(s)) - "
                      f"ainda restam {len(obs_restantes)}, tentando com teto {teto - 1} (mais conservador)...")
            else:
                print(f"  ⬇️  [{ws.title}] {len(obs_restantes)} OB(s) ainda sem match com teto de {teto} pagamento(s) "
                      f"- tentando com teto {teto - 1} (mais conservador)...")

    _pintar_grupos(ws, grupos_totais, col_ob, col_valor)

    return {
        "aba": ws.title,
        "grupos": grupos_totais,
        "obs_sem_match": obs_restantes,
        "pagamentos_sem_ob": pagamentos_restantes,
        "pagamentos_todos": pagamentos,
        "total_obs": len(obs),
        "max_combinacao_usado": teto_usado,
    }