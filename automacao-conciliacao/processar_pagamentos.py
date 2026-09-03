"""
Módulo de constantes/helpers compartilhados (fonte, cor, borda, localização
de blocos na planilha) usado por remover_zero.py, separar_tipo.py,
conciliacao.py e extratos.py.
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FONT_NAME = "Arial"
NORMAL_FONT = Font(name=FONT_NAME, size=10)
BOLD_FONT = Font(name=FONT_NAME, size=10, bold=True)
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NO_BORDER = Border()
NO_FILL = PatternFill(fill_type=None)
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

NUM_COLS_PRINCIPAL = 7  # A:G -> Linha, Tipo, Código, Data Transação, Quantia, Status, Justificativas

# Cores usadas para marcar cada grupo (pagamento(s) <-> OB). Se houver mais
# grupos que cores, elas se repetem.
CORES_CONCILIACAO = [
    "FFD966",  # amarelo forte
    "93C47D",  # verde
    "6FA8DC",  # azul
    "E06666",  # vermelho
    "C27BA0",  # rosa/magenta
    "8E7CC3",  # roxo
    "F6B26B",  # laranja
    "76A5AF",  # verde-azulado (teal)
    "A2C4C9",  # azul petróleo claro
    "FF9999",  # salmão
    "B4A7D6",  # lilás
    "FFE599",  # amarelo claro (variação, ainda contrastante)
    "45818E",  # teal escuro
    "CC4125",  # vermelho tijolo
    "674EA7",  # roxo escuro
    "38761D",  # verde escuro
    "0B5394",  # azul escuro
    "BF9000",  # dourado/mostarda
    "D5A6BD",  # rosa antigo
    "999999",  # cinza (reserva pra quando esgotar as outras)
    "6D9EEB",  # azul claro
    "8FCE00",  # verde limão
    "E69138",  # laranja queimado
    "A64D79",  # vinho/magenta escuro
    "16537E",  # azul petróleo escuro
    "B45F06",  # marrom alaranjado
    "783F04",  # marrom escuro
    "76D7C4",  # verde-água claro
    "F9CB9C",  # pêssego claro
    "D9D2E9",  # lilás muito claro
    "D0E0E3",  # azul gelo
    "EAD1DC",  # rosa bebê
    "FCE5CD",  # creme alaranjado
    "C9DAF8",  # azul bebê
    "D9EAD3",  # verde bem claro
    "FFF2CC",  # amarelo bem claro
    "B6D7A8",  # verde pastel
    "A4C2F4",  # azul pastel
    "F4CCCC",  # vermelho bem claro
]

def localizar_titulo(ws, prefixo):
    """Procura, na coluna A, a primeira linha cujo valor comece com
    `prefixo` (ex.: 'PAGAMENTOS', que aparece como 'PAGAMENTOS (47)').
    Devolve o número da linha ou None."""
    prefixo = prefixo.strip().upper()
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v and str(v).strip().upper().startswith(prefixo):
            return r
    return None


def localizar_bloco_ob(ws):
    """Procura em toda a aba uma célula 'OB' seguida, na coluna seguinte e
    mesma linha, por uma célula 'VALOR'. Devolve
    (col_ob, col_valor, linha_cabecalho) ou (None, None, None) se a aba não
    tiver essa tabela auxiliar."""
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v and str(v).strip().upper() == "OB":
                v_valor = ws.cell(row=r, column=c + 1).value
                if v_valor and str(v_valor).strip().upper() == "VALOR":
                    return c, c + 1, r
    return None, None, None


def _ler_linhas_ob(ws, col_ob, col_valor, linha_inicial):
    """Lê (ob, valor) a partir de `linha_inicial`, parando na primeira linha
    totalmente vazia nas duas colunas, ou na primeira linha em que a coluna
    VALOR contenha algo que não seja número (sinal de que a tabela OB/VALOR
    terminou e começou outra tabela colada em seguida, sem linha vazia de
    separação)."""
    linhas = []
    r = linha_inicial
    while True:
        v_ob = ws.cell(row=r, column=col_ob).value
        v_valor = ws.cell(row=r, column=col_valor).value
        if v_ob is None and v_valor is None:
            break
        if v_valor is not None and not isinstance(v_valor, (int, float)):
            break
        linhas.append((v_ob, v_valor))
        r += 1
    return linhas


def _normalizar_abas(abas):
    """Normaliza a entrada de abas para uma lista de nomes."""
    if abas is None:
        return None
    if isinstance(abas, str):
        return [abas]
    return list(abas)


def _centavos(valor):
    """Converte pra inteiro em centavos, pra comparar sem erro de ponto flutuante."""
    return round(float(valor) * 100)